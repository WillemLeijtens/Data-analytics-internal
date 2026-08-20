"""End-to-end: van onbekend bestand naar werkende retailer-analyses.

Parsers worden in het project gebouwd, niet in de app — zelf mappen is er
bewust uit. De keten voor een nieuwe retailer:
  1. onbekend bestand uploaden  -> PROFIEL NODIG, kolommen gesnifft (die
     informatie is het startpunt voor de parserbouw in het project)
  2. het project levert een profiel (hier: save_profile, in het echt een
     JSON in profiles/ of een ingebouwde parser)
  3. bestand opnieuw uploaden   -> herkend en ingelezen
  4. meerjarige historie laden  -> dashboard/analyses tonen de juiste cijfers
"""

import importlib
import io
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from engine import parser as parser_mod  # noqa: E402


def make_xlsx(headers, rows, sheet="Sheet1", meta_rows=0) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i in range(meta_rows):
        ws.cell(row=i + 1, column=1, value=f"metadata {i + 1}")
    for c, h in enumerate(headers, start=1):
        ws.cell(row=meta_rows + 1, column=c, value=h)
    for r, row in enumerate(rows, start=meta_rows + 2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


DG_HEADERS = ["Kalenderwoche", "Artikelnummer", "Marke", "Absatz", "Umsatz"]


def douglas_definition(break_period=False) -> dict:
    return {
        "detection": {"filename_glob": "Douglas_Abverkauf_KW*.xlsx", "sheet": "Sheet1",
                      "header_row": 1, "required_headers": DG_HEADERS[:3],
                      "filetype": "xlsx", "csv_delimiter": None, "decimal": ","},
        "period": {"type": "week",
                   "source_column": "BestaatNiet" if break_period else "Kalenderwoche",
                   "format": "yyyy-Www"},
        "mapping": [{"source": "Marke", "target": "merk"},
                    {"source": "Absatz", "target": "volume"},
                    {"source": "Umsatz", "target": "omzet"}],
        "constants": {"land": "DE"},
        "thresholds": {"promo_price_drop": 0.05},
    }


def ship_profile(retailer_id: str, definition: dict, status: str = "live"):
    """Zoals het project een profiel meelevert: rechtstreeks in de database,
    niet via een endpoint — publiceren kan niet meer vanuit de app."""
    import db
    from engine.profile import save_profile
    with db.get_conn() as conn:
        p = save_profile(conn, retailer_id, definition, status)
        if status == "live":
            conn.execute("UPDATE retailers SET aangesloten=1 WHERE id=?", (retailer_id,))
        return p


def dg_rows(year, weeks, factor=1.0):
    out = []
    for wk in weeks:
        out.append([f"{year}-W{wk:02d}", "DG-88231", "TWEEZERMAN", 400 + wk, round((400 + wk) * 17.5 * factor, 2)])
        out.append([f"{year}-W{wk:02d}", "DG-91002", "ALESSANDRO", 180 + wk, round((180 + wk) * 14.0 * factor, 2)])
    return out


# ------------------------------------------------------------------ sniff

def test_sniff_csv_metadata_and_comments():
    content = ("# commentaarregel\nExportdatum;2026-08-14\n"
               "Week;EAN;Merk;Stuks;Omzet\n2026-W32;4049469072773;X;10;120.50\n").encode()
    s = parser_mod.sniff("rapport.csv", content)
    assert s and s["filetype"] == "csv" and s["csv_delimiter"] == ";"
    assert s["columns"] == ["Week", "EAN", "Merk", "Stuks", "Omzet"]
    assert s["voorbeeld"][0] == "2026-W32"


def test_sniff_xlsx_with_metadata_block():
    content = make_xlsx(DG_HEADERS, dg_rows(2026, [32]), sheet="Data", meta_rows=6)
    s = parser_mod.sniff("rapport.xlsx", content)
    assert s and s["filetype"] == "xlsx" and s["sheet"] == "Data"
    assert s["header_row"] == 7 and s["columns"] == DG_HEADERS


def test_sniff_garbage_returns_none():
    assert parser_mod.sniff("kapot.xlsx", b"dit is geen bestand") is None


# ------------------------------------------------------------------ e2e

@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("CONSOLE_DB", str(tmp_path / "console.db"))
    monkeypatch.setenv("CONSOLE_AUTH", "gateway")
    monkeypatch.setenv("CONSOLE_BIND", "127.0.0.1")
    for name in ("db", "seed", "main"):
        sys.modules.pop(name, None)
    main = importlib.import_module("main")
    return TestClient(main.app)


def upload(client, filename, content):
    r = client.post("/api/import", files=[("files", (filename, content))])
    assert r.status_code == 200
    return r.json()["results"][0]


def test_full_parser_flow(client):
    # 1. Unknown file -> PROFIEL NODIG; the sniffed columns are the input
    #    for building the parser in the project.
    f2026 = make_xlsx(DG_HEADERS, dg_rows(2026, [30, 31, 32]))
    result = upload(client, "Douglas_Abverkauf_KW32.xlsx", f2026)
    assert result["status"] == "profiel_nodig"
    assert result["sniff"]["columns"] == DG_HEADERS

    # 2. The project ships a profile (v1 live).
    ship_profile("douglas", douglas_definition())
    profs = [p for p in client.get("/api/parser/profielen").json()
             if p["retailer_id"] == "douglas"]
    assert [(p["version"], p["status"]) for p in profs] == [(1, "live")]

    # Een onvolledig profiel meeleveren blijft geweigerd worden.
    bad = douglas_definition()
    bad["period"] = {"type": "week", "source_column": "", "format": "yyyy-Www"}
    with pytest.raises(ValueError):
        ship_profile("douglas", bad)

    # 3. Same file again -> recognised and loaded (replaces the old row)
    result = upload(client, "Douglas_Abverkauf_KW32.xlsx", f2026)
    assert result["status"] == "ingelezen" and result["rows"] == 6
    imports = client.get("/api/imports").json()
    assert [i["status"] for i in imports] == ["ingelezen"]

    # 4. History: previous year through the same pipeline, cheaper week 31
    f2025 = make_xlsx(DG_HEADERS, dg_rows(2025, [30, 31, 32], factor=0.9))
    result = upload(client, "Douglas_Abverkauf_KW31.xlsx", f2025)
    assert result["status"] == "ingelezen" and result["rows"] == 6

    dash = client.get("/api/douglas/dashboard").json()
    assert dash["available"] and not dash["empty"]
    assert dash["laatste_periode"] == "2026-W32"
    assert dash["trend"]["jaren"] == [2025, 2026]
    # YTD 2026 (w30-32) vs 2025 (w30-32): factor 0.9 -> ~ +11.1%
    assert dash["ytd"]["omzet"]["delta_pct"] == pytest.approx(11.1, abs=0.2)
    # Both brands separated, per the row's own Marke column
    merken = {b["merk"] for b in dash["kpi"]["omzet"]["breakdown"]}
    assert merken == {"TWEEZERMAN", "ALESSANDRO"}
    # No store data -> per-store KPI is an estimate (labels say so)
    assert "SCHATTING" in dash["labels"]

    # Article-level analyses stay off with a visible reason (no artikel_ean)
    art = client.get("/api/douglas/artikelen").json()
    assert not art["available"] and "OP MERKNIVEAU" in art["labels"]

    # Promotions run on merk+land scope without banner
    promo = client.get("/api/douglas/promoties").json()
    assert promo["available"]
    assert promo["resolution"]["level_used"]["scope"] == "merk+land"

    # Overview now shows douglas as connected with a live profile
    ov = client.get("/api/overview").json()
    dg = next(c for c in ov["retailers"] if c["id"] == "douglas")
    assert dg["aangesloten"] and dg["profiel"]["status"] == "live"


def test_databases_survive_restart(client, tmp_path, monkeypatch):
    """Profiles and facts live in the mounted database file, not in the
    container: a fresh app process on the same file sees everything."""
    f = make_xlsx(DG_HEADERS, dg_rows(2026, [32]))
    ship_profile("douglas", douglas_definition())
    upload(client, "Douglas_Abverkauf_KW32.xlsx", f)

    # Simulate a restart: re-import main against the SAME database file.
    for name in ("db", "seed", "main"):
        sys.modules.pop(name, None)
    main2 = importlib.import_module("main")
    c2 = TestClient(main2.app)
    profs = [p for p in c2.get("/api/parser/profielen").json() if p["retailer_id"] == "douglas"]
    assert profs[0]["status"] == "live"
    assert c2.get("/api/douglas/dashboard").json()["available"]
