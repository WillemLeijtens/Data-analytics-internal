"""Het echte ICI Paris XL-maandrapport + de retailer-scheiding."""

import importlib
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from test_parser_flow import upload  # noqa: E402

U = Path("/root/.claude/uploads/54377bab-ac94-5cbf-8750-c3a4d90899e0")
REAL_ICI = U / "fc6fc987-Maandelijkse_resultaten__Tweezerman__Depend_ICI_Paris_XL__4.xlsx"
REAL_KV = U / "d62acf54-DWH__Sales_volume__sales_Tweezerman_KVNL_1299_1734396111539283577.xlsx"


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("CONSOLE_DB", str(tmp_path / "console.db"))
    monkeypatch.setenv("CONSOLE_AUTH", "gateway")
    monkeypatch.setenv("CONSOLE_BIND", "127.0.0.1")
    for name in ("db", "seed", "main"):
        sys.modules.pop(name, None)
    main = importlib.import_module("main")
    return TestClient(main.app)


def test_generated_ici_report_imports(client):
    import seed
    content = seed.make_ici_xlsx(seed._ici_demo_blocks(["202606", "202607"]))
    r = upload(client, "Maandelijkse_resultaten__Tweezerman__Depend_ICI_Paris_XL__x.xlsx", content)
    assert r["status"] == "ingelezen" and r["retailer_id"] == "ici-paris-xl"
    assert r["rows"] == 12          # 2 merken x 3 winkels x 2 maanden

    dash = client.get("/api/ici-paris-xl/dashboard").json()
    assert dash["periode_type"] == "maand"
    assert dash["laatste_periode"] == "2026-07"
    assert "OP MAANDNIVEAU" in dash["labels"]
    assert dash["capabilities"]["volume"] is False
    # Winkelaantal komt uit de feiten, geen SCHATTING.
    assert dash["kpi"]["omzet_per_winkel"]["schatting"] is False
    assert dash["kpi"]["omzet_per_winkel"]["winkels"] == 3

    promo = client.get("/api/ici-paris-xl/promoties").json()
    # Geen volume -> geen suggesties, maar wel elke periode handmatig aan te vinken.
    assert all(s["suggestie"] is None for s in promo["suggesties"])
    assert len(promo["suggesties"]) == 4    # 2 merken x 2 maanden


@pytest.mark.skipif(not (REAL_ICI.exists() and REAL_KV.exists()),
                    reason="echte sample-bestanden niet aanwezig")
def test_both_real_files_one_batch_no_cross_contamination(client):
    """De kernvraag: meerdere bestanden tegelijk, elk bij de juiste retailer,
    zonder dat ICI-data bij Kruidvat belandt of andersom."""
    r = client.post("/api/import", files=[
        ("files", (REAL_ICI.name[9:], REAL_ICI.read_bytes())),
        ("files", (REAL_KV.name[9:], REAL_KV.read_bytes())),
    ]).json()["results"]
    by_name = {x["filename"]: x for x in r}
    assert by_name[REAL_ICI.name[9:]]["retailer_id"] == "ici-paris-xl"
    assert by_name[REAL_ICI.name[9:]]["status"] == "ingelezen"
    assert by_name[REAL_KV.name[9:]]["retailer_id"] == "kruidvat"
    assert by_name[REAL_KV.name[9:]]["status"] == "ingelezen"

    import db
    with db.get_conn() as conn:
        rows = conn.execute(
            "SELECT retailer_id, periode_type, COUNT(*) n, COUNT(DISTINCT merk) merken "
            "FROM sellout_facts GROUP BY retailer_id").fetchall()
        per = {r["retailer_id"]: r for r in rows}
        # ICI: maandfeiten met winkel; Kruidvat: weekfeiten zonder winkel.
        assert per["ici-paris-xl"]["periode_type"] == "maand"
        assert per["kruidvat"]["periode_type"] == "week"
        cross = conn.execute(
            "SELECT COUNT(*) c FROM sellout_facts WHERE "
            "(retailer_id='kruidvat' AND winkel_id IS NOT NULL) OR "
            "(retailer_id='ici-paris-xl' AND winkel_id IS NULL)").fetchone()["c"]
        assert cross == 0

    # Beide dashboards tonen hun eigen wereld.
    ici = client.get("/api/ici-paris-xl/dashboard").json()
    kv = client.get("/api/kruidvat/dashboard").json()
    assert ici["laatste_periode"].count("-W") == 0
    assert "-W" in kv["laatste_periode"]
    assert {b["merk"] for b in ici["kpi"]["omzet"]["breakdown"]} == {"TWEEZERMAN", "DEPEND"}


@pytest.mark.skipif(not REAL_ICI.exists(), reason="echt sample-bestand niet aanwezig")
def test_real_ici_reconciles_with_its_own_brand_tab(client):
    r = upload(client, "hernoemd-ici-rapport (kopie).xlsx", REAL_ICI.read_bytes())
    assert r["status"] == "ingelezen" and r["retailer_id"] == "ici-paris-xl"
    assert r["detail"] is None, f"reconciliatie-waarschuwing: {r['detail']}"
    assert r["rows"] == 4355


def test_ici_missing_reconciliation_tab_fails_closed(client):
    import io
    import seed
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(seed.make_ici_xlsx(seed._ici_demo_blocks(["202607"]))))
    del wb["Brands"]
    out = io.BytesIO()
    wb.save(out)
    r = upload(client, "Maandelijkse_resultaten_ICI_Paris_geen-brands.xlsx", out.getvalue())
    assert r["status"] == "error" and "reconciliatie niet mogelijk" in r["detail"]


def test_ici_missing_scope_is_not_silently_approved(client):
    """Een merk dat in de winkeltabs ontbreekt maar wél in de merk-tab staat,
    mag niet door de reconciliatie glippen."""
    import io
    import seed
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(seed.make_ici_xlsx(seed._ici_demo_blocks(["202607"]))))
    wb["Brands"].cell(row=3, column=3, value="ONTBREKEND_MERK")
    out = io.BytesIO()
    wb.save(out)
    r = upload(client, "Maandelijkse_resultaten_ICI_Paris_ontbrekend.xlsx", out.getvalue())
    assert r["status"] == "error" and "merk/maand" in r["detail"]


def _blocks_twee_jaar():
    """202512 = vorig jaar, 202601 t/m 202603 = huidig jaar.

    TWEEZERMAN: 6051 draait door, 6052 ligt twee maanden stil (gestopt),
    6053 is nieuw, 6054 mist alleen de laatste maand (signaal).
    DEPEND: alleen 6051 — een kleiner winkelbestand dan TWEEZERMAN.
    """
    return {
        "TWEEZERMAN": {
            "6051": {"202512": 100.0, "202601": 200.0, "202602": 250.0, "202603": 300.0},
            "6052": {"202512": 900.0, "202601": 50.0},
            "6053": {"202601": 40.0, "202602": 50.0, "202603": 60.0},
            "6054": {"202512": 500.0, "202601": 10.0, "202602": 20.0},
        },
        "DEPEND": {"6051": {"202512": 10.0, "202601": 20.0,
                            "202602": 25.0, "202603": 30.0}},
    }


def test_winkelaantal_telt_alleen_winkels_met_omzet_dit_jaar(client):
    """Per merk gedeeld door de winkels die dít jaar omzet gaven — niet door
    de hele winkellijst en niet door alleen de verkopers van die maand."""
    import seed
    upload(client, "Maandelijkse_resultaten_ICI_Paris_XL_2j.xlsx",
           seed.make_ici_xlsx(_blocks_twee_jaar()))
    k = client.get("/api/ici-paris-xl/dashboard").json()["kpi"]["omzet_per_winkel"]
    per_merk = {b["merk"]: b for b in k["breakdown"]}
    # TWEEZERMAN: alle vier de winkels hadden in 2026 omzet, ook 6052 en 6054
    # die de laatste maand(en) niets meer verkochten.
    assert per_merk["TWEEZERMAN"]["winkels"] == 4
    assert per_merk["TWEEZERMAN"]["waarde"] == pytest.approx(360.0 / 4)
    # DEPEND heeft één winkel; delen door 4 zou het merk vier keer te laag zetten.
    assert per_merk["DEPEND"]["winkels"] == 1
    assert per_merk["DEPEND"]["waarde"] == pytest.approx(30.0)
    assert all(b["schatting"] is False for b in k["breakdown"])


def test_winkelanalyse_gestopt_en_toegevoegd(client):
    import seed
    upload(client, "Maandelijkse_resultaten_ICI_Paris_XL_2j.xlsx",
           seed.make_ici_xlsx(_blocks_twee_jaar()))
    w = client.get("/api/ici-paris-xl/dashboard").json()["winkelanalyse"]
    assert w["beschikbaar"] is True and w["jaar"] == 2026 and w["laatste_maand"] == 3
    assert "category manager" in w["actiepunt"] and w["gestopt_vanaf"] == 2

    # Gestopt: pas vanaf twee lege maanden.
    gestopt = {(g["winkel_id"], g["merk"]): g for g in w["gestopt"]}
    assert set(gestopt) == {("6052", "TWEEZERMAN")}
    g = gestopt[("6052", "TWEEZERMAN")]
    assert g["laatste_maand"] == 1 and g["maanden_zonder_omzet"] == 2
    # De omzet van vorig jaar is wat we nu mislopen.
    assert g["omzet_vorig_jaar"] == pytest.approx(900.0)
    assert w["gemiste_omzet"] == pytest.approx(900.0)

    # Eén lege maand telt niet als gestopt, maar verdwijnt ook niet: 6054
    # staat als signaal in de lijst en niet in de gemiste omzet.
    signalen = {(s["winkel_id"], s["merk"]): s for s in w["signalen"]}
    assert set(signalen) == {("6054", "TWEEZERMAN")}
    assert signalen[("6054", "TWEEZERMAN")]["maanden_zonder_omzet"] == 1
    assert signalen[("6054", "TWEEZERMAN")]["omzet_vorig_jaar"] == pytest.approx(500.0)

    toegevoegd = {(a["winkel_id"], a["merk"]) for a in w["toegevoegd"]}
    assert toegevoegd == {("6053", "TWEEZERMAN")}
    assert w["toegevoegd"][0]["eerste_maand"] == 1
    assert w["toegevoegd"][0]["omzet_dit_jaar"] == pytest.approx(150.0)


def test_winkelanalyse_ontbreekt_zonder_winkeldata(client):
    """Kruidvat levert geen winkelniveau — dan hoort er geen winkelanalyse
    te staan in plaats van een lege of verzonnen lijst."""
    import seed
    upload(client, "DWH__Sales_volume__sales_Tweezerman_KVNL_32.xlsx",
           seed.make_dwh_xlsx(seed._kv_demo_rows(["202632"])))
    dash = client.get("/api/kruidvat/dashboard").json()
    assert dash["winkelanalyse"] == {"beschikbaar": False}


def test_controle_meldt_retailer_zonder_te_importeren(client):
    """De controlestap kijkt in het bestand en meldt de retailer, maar mag
    nog niets opslaan — pas na bevestiging wordt er geïmporteerd."""
    import seed
    content = seed.make_ici_xlsx(seed._ici_demo_blocks(["202607"]))

    r = client.post("/api/import/controle", files=[
        ("files", ("Maandelijkse resultaten ICI Paris XL (4).xlsx", content))]).json()["results"][0]
    assert r["herkend"] is True
    assert r["retailer_id"] == "ici-paris-xl"
    assert r["retailer_naam"] == "ICI Paris XL"
    assert client.get("/api/imports").json() == [], "controle mag niets opslaan"

    # Pas na bevestiging landt het bestand in de database.
    upload(client, "Maandelijkse resultaten ICI Paris XL (4).xlsx", content)
    assert len(client.get("/api/imports").json()) == 1


def test_controle_meldt_onbekend_formaat(client):
    r = client.post("/api/import/controle", files=[
        ("files", ("onbekend_rapport.xlsx", b"geen spreadsheet"))]).json()["results"][0]
    assert r["herkend"] is False and r["retailer_id"] is None
    assert "geen parser" in r["detail"].lower()
    assert client.get("/api/imports").json() == []
