"""De echte Etos Data Grid-widgetexport + de fail-closed-controles.

Dit formaat heeft géén totalenrij; de parser verifieert daarom alles wat het
bestand zelf aan controle-informatie biedt (merkental, weekbereik en de
ISO-einddatums uit het metadatablok) en weigert bij elke afwijking.
"""

import importlib
import io
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from test_parser_flow import upload  # noqa: E402

U = Path("/root/.claude/uploads/54377bab-ac94-5cbf-8750-c3a4d90899e0")
REAL_ETOS = U / "b9ccb189-Data_Grid_57018_widget.xlsx"
REAL_KV = U / "d62acf54-DWH__Sales_volume__sales_Tweezerman_KVNL_1299_1734396111539283577.xlsx"
REAL_ICI = U / "fc6fc987-Maandelijkse_resultaten__Tweezerman__Depend_ICI_Paris_XL__4.xlsx"

# Historiebestanden met andere Time-scopes (kwartalen, twee kwartalen,
# fiscale periodes). Verwachte cijfers komen uit een ONAFHANKELIJKE telling
# buiten de parser om: som van alle gevulde Sales/Units-cellen per bestand.
REAL_HISTORIE = [
    ("bd12efd7-Data_Grid_57018_widget_Q1_2025.xlsx",
     438, 334302.90, 17220, "2025-W01", "2025-W16"),
    ("728a9bd5-Data_Grid_57018_widget_Q2enQ3_2025.xlsx",
     682, 685636.53, 40469, "2025-W17", "2025-W40"),
    ("7fe006b1-Data_Grid_57018_widget_Q4_2025.xlsx",
     328, 342033.51, 21518, "2025-W41", "2025-W52"),
    ("c62214ae-Data_Grid_57018_widget_2024.xlsx",
     685, 638163.13, 36463, "2024-W19", "2024-W52"),
]
HISTORIE_AANWEZIG = all((U / naam).exists() for naam, *_ in REAL_HISTORIE)


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("CONSOLE_DB", str(tmp_path / "console.db"))
    monkeypatch.setenv("CONSOLE_AUTH", "gateway")
    monkeypatch.setenv("CONSOLE_BIND", "127.0.0.1")
    for name in ("db", "seed", "main"):
        sys.modules.pop(name, None)
    main = importlib.import_module("main")
    return TestClient(main.app)


def _demo(weeks=("202630", "202631", "202632"), **kw):
    import seed
    return seed.make_etos_xlsx(seed._etos_demo_artikelen(list(weeks)), **kw)


def _herbouw(content: bytes, muteer) -> bytes:
    """Laad het gegenereerde bestand, pas het aan, geef de bytes terug."""
    wb = load_workbook(io.BytesIO(content))
    muteer(wb[wb.sheetnames[0]])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------- echt bestand

@pytest.mark.skipif(not REAL_ETOS.exists(), reason="echt sample-bestand niet aanwezig")
def test_real_etos_full_counts(client):
    """Alle cijfers uit het echte bestand, gepind tegen een onafhankelijke
    telling (som van alle gevulde Sales/Units-cellen, buiten de parser om):
    1126 gevulde weekparen, EUR 900.697,04, 61.313 stuks."""
    r = upload(client, REAL_ETOS.name[9:], REAL_ETOS.read_bytes())
    assert r["status"] == "ingelezen" and r["retailer_id"] == "etos"
    assert r["rows"] == 1126

    import db
    with db.get_conn() as conn:
        tot = conn.execute(
            "SELECT SUM(omzet) o, SUM(volume) v, COUNT(DISTINCT artikel_ean) a, "
            "COUNT(DISTINCT merk) m FROM sellout_facts WHERE retailer_id='etos'").fetchone()
        assert tot["o"] == pytest.approx(900697.04, abs=0.01)
        assert tot["v"] == 61313
        assert tot["a"] == 49 and tot["m"] == 3
        merken = {r2["merk"] for r2 in conn.execute(
            "SELECT DISTINCT merk FROM sellout_facts WHERE retailer_id='etos'")}
        assert merken == {"BJÖRN AXÉN", "PATCHOLOGY", "TWEEZERMAN"}

    dash = client.get("/api/etos/dashboard").json()
    assert dash["periode_type"] == "week"
    assert dash["laatste_periode"] == "2026-W32"
    assert dash["capabilities"]["volume"] is True
    assert dash["capabilities"]["artikel"] is True
    # Geen 2025 in dit YTD-bestand: geen delta, wél gemeld als niet-vergelijkbaar.
    assert dash["ytd"]["omzet"]["delta_pct"] is None
    assert set(dash["ytd"]["basis"]["niet_vergelijkbaar"]) == \
        {"BJÖRN AXÉN", "PATCHOLOGY", "TWEEZERMAN"}

    art = client.get("/api/etos/artikelen").json()
    assert art["available"] and len(art["artikelen"]) == 49
    promo = client.get("/api/etos/promoties").json()
    assert promo["methode"] == "prijsindex"


@pytest.mark.skipif(not REAL_ETOS.exists(), reason="echt sample-bestand niet aanwezig")
def test_real_etos_renamed_still_recognised(client):
    r = upload(client, "kopie van rapportage etos (3).xlsx", REAL_ETOS.read_bytes())
    assert r["status"] == "ingelezen" and r["retailer_id"] == "etos"


@pytest.mark.skipif(not (REAL_ETOS.exists() and REAL_KV.exists() and REAL_ICI.exists()),
                    reason="echte sample-bestanden niet aanwezig")
def test_three_real_retailers_one_batch(client):
    r = client.post("/api/import", files=[
        ("files", (REAL_ETOS.name[9:], REAL_ETOS.read_bytes())),
        ("files", (REAL_KV.name[9:], REAL_KV.read_bytes())),
        ("files", (REAL_ICI.name[9:], REAL_ICI.read_bytes())),
    ]).json()["results"]
    per = {x["retailer_id"]: x["status"] for x in r}
    assert per == {"etos": "ingelezen", "kruidvat": "ingelezen",
                   "ici-paris-xl": "ingelezen"}
    import db
    with db.get_conn() as conn:
        cross = conn.execute(
            "SELECT COUNT(*) c FROM sellout_facts WHERE "
            "(retailer_id='etos' AND (winkel_id IS NOT NULL OR banner IS NOT NULL "
            " OR land != 'NL' OR artikel_ean IS NULL)) OR "
            "(retailer_id='etos' AND periode_type != 'week')").fetchone()["c"]
        assert cross == 0


# ---------------------------------------------------------------- historie

@pytest.mark.skipif(not HISTORIE_AANWEZIG, reason="historiebestanden niet aanwezig")
def test_real_historie_alle_scopes(client):
    """De vier echte historiebestanden gebruiken andere Time-scopes
    (Fiscal Quarter met weekrange, 2 Fiscal Quarters, 9 Fiscal Periods).
    Elk moet inlezen met exact de onafhankelijk getelde totalen."""
    import db
    for naam, n, omzet, volume, eerste, laatste in REAL_HISTORIE:
        r = upload(client, naam[9:], (U / naam).read_bytes())
        assert r["status"] == "ingelezen" and r["retailer_id"] == "etos", \
            f"{naam}: {r['status']} — {r.get('detail')}"
        assert r["rows"] == n, naam
    with db.get_conn() as conn:
        for naam, n, omzet, volume, eerste, laatste in REAL_HISTORIE:
            tot = conn.execute(
                "SELECT SUM(omzet) o, SUM(volume) v, COUNT(*) n FROM sellout_facts "
                "WHERE retailer_id='etos' AND periode BETWEEN ? AND ?",
                (eerste, laatste)).fetchone()
            assert tot["n"] == n, naam
            assert tot["o"] == pytest.approx(omzet, abs=0.01), naam
            assert tot["v"] == volume, naam


@pytest.mark.skipif(not (HISTORIE_AANWEZIG and REAL_ETOS.exists()),
                    reason="niet alle echte bestanden aanwezig")
def test_alle_vijf_bestanden_samen_geeft_meerjarige_analyse(client):
    """2024 + 3×2025 + YTD 2026 in één batch: geen dubbeltelling over de
    bestanden heen, en de jaar-op-jaar-vergelijking komt tot leven."""
    files = [("files", (naam[9:], (U / naam).read_bytes()))
             for naam, *_ in REAL_HISTORIE]
    files.append(("files", (REAL_ETOS.name[9:], REAL_ETOS.read_bytes())))
    rr = client.post("/api/import", files=files).json()["results"]
    assert all(x["status"] == "ingelezen" and x["retailer_id"] == "etos" for x in rr), rr

    import db
    with db.get_conn() as conn:
        tot = conn.execute("SELECT SUM(omzet) o, COUNT(*) n FROM sellout_facts "
                           "WHERE retailer_id='etos'").fetchone()
        # Som van de vijf onafhankelijke tellingen — niets dubbel, niets kwijt.
        verwacht = 334302.90 + 685636.53 + 342033.51 + 638163.13 + 900697.04
        assert tot["o"] == pytest.approx(verwacht, abs=0.05)
        assert tot["n"] == 438 + 682 + 328 + 685 + 1126

    dash = client.get("/api/etos/dashboard").json()
    assert dash["trend"]["jaren"] == [2024, 2025, 2026]
    # 2025 is compleet (wk 1-52), dus de YoY 2026-vs-2025 heeft nu een echte
    # delta op volledige basis: alle drie de merken in beide jaren.
    assert dash["ytd"]["basis"]["volledig"] is True
    assert dash["ytd"]["omzet"]["delta_pct"] is not None
    assert dash["ytd"]["omzet"]["vorig"] > 0


# ---------------------------------------------------------------- gegenereerd

def test_generated_etos_imports_and_analyses(client):
    r = upload(client, "Data_Grid_99999_widget.xlsx", _demo())
    assert r["status"] == "ingelezen" and r["retailer_id"] == "etos"
    assert r["rows"] == 9      # 3 artikelen x 3 weken

    dash = client.get("/api/etos/dashboard").json()
    assert dash["laatste_periode"] == "2026-W32"
    # Weekomzet = som(units x prijs) — spotcheck tegen de generator.
    import seed
    verwacht = sum(s for a in seed._etos_demo_artikelen(["202630", "202631", "202632"])
                   for wk, (s, _u) in a["weeks"].items() if wk == "202632")
    assert dash["kpi"]["omzet"]["waarde"] == pytest.approx(verwacht)

    # Assortiment draait op UPC-niveau met rotatietarget.
    client.put("/api/etos/instellingen", json={
        "winkels_targets": [{"merk": "TWEEZERMAN", "land": "NL", "banner": None,
                             "aantal_winkels": 500, "target_per_winkel": None}],
        "rotatie_targets": [{"merk": "TWEEZERMAN", "stuks_per_winkel_per_week": 0.5}]})
    a = client.get("/api/etos/assortiment").json()
    assert a["available"] and len(a["artikelen"]) == 3


def test_herimport_ytd_vervangt_zonder_dubbeltelling(client):
    """Elke download is een groeiend YTD-bestand: wk 1..N+1 overlapt wk 1..N
    volledig. De natuurlijke sleutel moet vervangen, nooit optellen."""
    upload(client, "Data_Grid_1_widget.xlsx", _demo(("202630", "202631")))
    eerste = client.get("/api/etos/dashboard").json()["kpi"]["omzet"]["waarde"]
    upload(client, "Data_Grid_2_widget.xlsx", _demo(("202630", "202631", "202632")))
    dash = client.get("/api/etos/dashboard").json()
    assert dash["laatste_periode"] == "2026-W32"
    import db
    with db.get_conn() as conn:
        n = conn.execute("SELECT COUNT(*) c FROM sellout_facts WHERE retailer_id='etos' "
                         "AND periode='2026-W30'").fetchone()["c"]
        assert n == 3, "week 30 moet vervangen zijn, niet verdubbeld"
    assert eerste > 0


# ---------------------------------------------------------------- fail-closed

def test_merkental_mismatch_fails(client):
    r = upload(client, "Data_Grid_3_widget.xlsx", _demo(brand_count=4))
    assert r["status"] == "error" and "Brand (4)" in r["detail"]


def test_ontbrekende_weekkolom_fails(client):
    def weg(ws):
        # verwijder de kop van week 202631 -> gat in de range
        for cell in ws[20]:
            if cell.value and str(cell.value).startswith("202631"):
                cell.value = None
    r = upload(client, "Data_Grid_4_widget.xlsx", _herbouw(_demo(), weg))
    assert r["status"] == "error" and "202631" in r["detail"]


def test_ending_datum_niet_op_iso_zondag_fails(client):
    def schuif(ws):
        for cell in ws[20]:
            if cell.value and str(cell.value).startswith("202631"):
                cell.value = "202631 (Ending 03/08/2026)"   # maandag i.p.v. zondag
    r = upload(client, "Data_Grid_5_widget.xlsx", _herbouw(_demo(), schuif))
    assert r["status"] == "error" and "ISO-zondag" in r["detail"]


def test_dubbele_upc_week_fails(client):
    import seed
    art = seed._etos_demo_artikelen(["202632"])
    art.append(dict(art[0]))                      # zelfde UPC nogmaals
    r = upload(client, "Data_Grid_6_widget.xlsx", seed.make_etos_xlsx(art))
    assert r["status"] == "error" and "dubbel" in r["detail"].lower()


def test_brand_metadata_ontbreekt_fails(client):
    def wis(ws):
        # beide Brand (N)-vermeldingen weg
        ws.cell(row=11, column=2).value = None
        ws.cell(row=19, column=1).value = "Merken"
    r = upload(client, "Data_Grid_7_widget.xlsx", _herbouw(_demo(), wis))
    assert r["status"] == "error" and "Brand (N)" in r["detail"]


def test_scope_zonder_weekrange_wordt_geaccepteerd(client):
    """De kwartaal/periode-exports noemen geen weekrange; dan geldt de
    intrinsieke aaneengesloten-reeks-eis plus de Ending-kruiscontrole."""
    r = upload(client, "Data_Grid_8_widget.xlsx", _demo(scope="ending"))
    assert r["status"] == "ingelezen" and r["rows"] == 9
    r = upload(client, "Data_Grid_9_widget.xlsx", _demo(scope="weeks"))
    assert r["status"] == "ingelezen"


def test_gat_in_weekreeks_zonder_range_fails(client):
    """Zonder expliciete range moet een gat in de reeks alsnog opvallen."""
    def knip(ws):
        # week 202631 volledig weg (kop + subkop + data) -> gat 30..32
        for rij in range(20, 26):
            for cell in ws[rij]:
                if cell.column in (6, 7):     # kolommen van week 2 (202631)
                    cell.value = None
    r = upload(client, "Data_Grid_10_widget.xlsx",
               _herbouw(_demo(scope="ending"), knip))
    assert r["status"] == "error" and "202631" in r["detail"]


def test_scope_ending_hoort_bij_laatste_week_fails(client):
    """Een Ending-datum in de Time-scope die niet bij de laatste weekkolom
    hoort: metadata en kolommen gaan dan niet over dezelfde periode."""
    def schuif(ws):
        ws.cell(row=13, column=2).value = \
            'Time "2 Fiscal Quarters 202502-202503, Ending 26/07/2026"'
    r = upload(client, "Data_Grid_11_widget.xlsx",
               _herbouw(_demo(scope="ending"), schuif))
    assert r["status"] == "error" and "zelfde periode" in r["detail"]


def test_weeks_range_die_niet_klopt_fails(client):
    def rek(ws):
        ws.cell(row=13, column=2).value = \
            'Fiscal Quarter "202503 (Weeks 202630-202635, Ending 30/08/2026)"'
    r = upload(client, "Data_Grid_12_widget.xlsx",
               _herbouw(_demo(scope="weeks"), rek))
    assert r["status"] == "error" and "ontbreekt" in r["detail"]


def test_healthz_toont_profielen(client):
    h = client.get("/healthz").json()
    assert h["status"] == "ok"
    assert h["profielen"].get("etos") == 1
    assert "kruidvat" in h["profielen"] and "ici-paris-xl" in h["profielen"]


# ---------------------------------------------------------------- bootstrap

def test_bootstrap_laadt_nieuw_profiel_op_bestaande_database(client, tmp_path):
    """De oude bootstrap draaide alleen bij een lege profieltabel; een nieuw
    meegeleverd profiel (zoals Etos) kwam dan nooit aan op een bestaande
    installatie. Nu draait hij idempotent bij elke start."""
    import db
    import seed as seed_mod
    with db.get_conn() as conn:
        conn.execute("DELETE FROM parser_profiles WHERE retailer_id='etos'")
        conn.execute("UPDATE retailers SET aangesloten=0 WHERE id='etos'")
    # Simuleer een herstart op dezelfde database: bootstrap opnieuw.
    seed_mod.bootstrap()
    with db.get_conn() as conn:
        rij = conn.execute("SELECT status FROM parser_profiles WHERE retailer_id='etos' "
                           "ORDER BY version DESC LIMIT 1").fetchone()
        assert rij and rij["status"] == "live"
        assert conn.execute("SELECT aangesloten FROM retailers WHERE id='etos'")\
            .fetchone()["aangesloten"] == 1
        # Bestaande profielen zijn niet aangeraakt (geen extra versies).
        n_kv = conn.execute("SELECT COUNT(*) c FROM parser_profiles "
                            "WHERE retailer_id='kruidvat'").fetchone()["c"]
    seed_mod.bootstrap()
    with db.get_conn() as conn:
        assert conn.execute("SELECT COUNT(*) c FROM parser_profiles "
                            "WHERE retailer_id='kruidvat'").fetchone()["c"] == n_kv
