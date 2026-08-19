"""Het eenmalige inleesscript voor de Kruidvat-export op winkelniveau.

Die export heeft een andere indeling dan de reguliere DWH-feed en wordt door
de parser afgewezen. tools/eenmalig_winkelbestand.py leest hem apart in. De
twee dingen die daarbij mis kunnen gaan staan hier vast:

  * dubbeltelling — de feitensleutel bevat winkel_id, dus winkelrijen komen
    NAAST bestaande weekregels te staan in plaats van eroverheen;
  * een tweede keer draaien dat de lading verdubbelt.
"""

from __future__ import annotations

import importlib
import io
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from test_parser_flow import upload  # noqa: E402

TOOLS = Path(__file__).resolve().parents[1] / "tools"


def maak_winkelbestand(rijen, weken, land="BEL", formule="KV") -> bytes:
    """Bouwt de echte indeling na: metadata verspreid over rij 1-20 met de
    waarde in kolom C (en bij Brand één rij bóven het label, zoals de
    samengevoegde cellen in het origineel), kop op rij 22/23, data vanaf 24,
    plus een Total-kolom en een tweede blad met dagcijfers.

    rijen: [{sku, merk, omschrijving, winkel, plaats, weken: {"202526": (vol, val)}}]
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales per week"
    ws.cell(row=2, column=2, value="Country:")
    ws.cell(row=2, column=3, value=land)
    ws.cell(row=3, column=2, value="Formula:")
    ws.cell(row=3, column=3, value=formule)
    # Waarde boven het label — precies de valkuil uit het echte bestand.
    ws.cell(row=11, column=3, value=";".join(sorted({r["merk"] for r in rijen})))
    ws.cell(row=12, column=2, value="Brand:")

    kop = ["SKU no.", "Brand", "Articledescription", "Size", "Type",
           "Package", "Store", "City", "Stock", "Stock"]
    for i, naam in enumerate(kop, start=2):
        ws.cell(row=22, column=i, value=naam)
    ws.cell(row=23, column=11, value="position/ pcs")
    ws.cell(row=23, column=12, value="Sales value")

    kolom = {}
    c = 13
    for w in weken:
        ws.cell(row=22, column=c, value=w)
        ws.cell(row=23, column=c, value="Sales volume")
        ws.cell(row=23, column=c + 1, value="Gross sales value")
        kolom[w] = c
        c += 2
    ws.cell(row=22, column=c, value="Total")
    ws.cell(row=23, column=c, value="Sales volume")

    for i, r in enumerate(rijen, start=24):
        ws.cell(row=i, column=2, value=float(r["sku"]))     # numeriek, net als echt
        ws.cell(row=i, column=3, value=r["merk"])
        ws.cell(row=i, column=4, value=r.get("omschrijving", "ARTIKEL"))
        ws.cell(row=i, column=8, value=float(r["winkel"]))
        ws.cell(row=i, column=9, value=r.get("plaats", "ANTWERPEN"))
        for w, (vol, val) in r["weken"].items():
            ws.cell(row=i, column=kolom[w], value=vol)
            ws.cell(row=i, column=kolom[w] + 1, value=val)

    dag = wb.create_sheet("Sales per day")
    dag.cell(row=22, column=2, value="SKU no.")
    dag.cell(row=22, column=7, value="Store")
    dag.cell(row=24, column=2, value=999999.0)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture()
def omgeving(tmp_path, monkeypatch):
    monkeypatch.setenv("CONSOLE_DB", str(tmp_path / "console.db"))
    monkeypatch.setenv("CONSOLE_AUTH", "gateway")
    monkeypatch.setenv("CONSOLE_BIND", "127.0.0.1")
    for naam in ("db", "seed", "main"):
        sys.modules.pop(naam, None)
    main = importlib.import_module("main")
    sys.path.insert(0, str(TOOLS))
    sys.modules.pop("eenmalig_winkelbestand", None)
    script = importlib.import_module("eenmalig_winkelbestand")
    return TestClient(main.app), script


def _bestand(tmp_path, inhoud: bytes) -> str:
    pad = tmp_path / "DWH__per_week_per_store_KVBE.xlsx"
    pad.write_bytes(inhoud)
    return str(pad)


def _twee_winkels(weken):
    return [
        {"sku": "6369354", "merk": "ALESSANDRO", "winkel": "8006", "plaats": "ANTWERPEN",
         "weken": {w: (2, 20.0) for w in weken}},
        {"sku": "6369354", "merk": "ALESSANDRO", "winkel": "8009", "plaats": "HACCOURT",
         "weken": {w: (3, 30.0) for w in weken}},
        {"sku": "6370075", "merk": "DEPEND GEL IQ", "winkel": "8006", "plaats": "ANTWERPEN",
         "weken": {w: (5, 50.0) for w in weken}},
    ]


def _feiten(client):
    import db
    with db.get_conn() as conn:
        return tuple(conn.execute(
            "SELECT COUNT(*), COALESCE(ROUND(SUM(omzet),2),0) FROM sellout_facts "
            "WHERE retailer_id='kruidvat'").fetchone())


def test_proefronde_schrijft_niets(omgeving, tmp_path):
    client, script = omgeving
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(["202540", "202541"]),
                                                ["202540", "202541"]))
    voor = _feiten(client)
    assert script.main([pad]) == 0
    assert _feiten(client) == voor == (0, 0)


def test_leest_winkels_weken_en_merken(omgeving, tmp_path):
    client, script = omgeving
    weken = ["202540", "202541"]
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(weken), weken))
    assert script.main([pad, "--schrijf"]) == 0

    import db
    with db.get_conn() as conn:
        rij = conn.execute(
            "SELECT COUNT(*), COUNT(DISTINCT winkel_id), COUNT(DISTINCT periode), "
            "ROUND(SUM(omzet),2), ROUND(SUM(volume),2) FROM sellout_facts "
            "WHERE retailer_id='kruidvat'").fetchone()
    # 3 regels x 2 weken; (20+30+50) per week = 100
    assert tuple(rij) == (6, 2, 2, 200.0, 20.0)
    with db.get_conn() as conn:
        namen = dict(conn.execute(
            "SELECT DISTINCT winkel_id, winkel_naam FROM sellout_facts "
            "WHERE retailer_id='kruidvat'").fetchall())
    # Winkelnummers komen als float uit Excel; '8006.0' mag niet in de sleutel.
    assert namen == {"8006": "ANTWERPEN", "8009": "HACCOURT"}


def test_dagblad_telt_niet_mee(omgeving, tmp_path):
    client, script = omgeving
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(["202540"]), ["202540"]))
    script.main([pad, "--schrijf"])
    import db
    with db.get_conn() as conn:
        skus = {r[0] for r in conn.execute(
            "SELECT DISTINCT artikel_ean FROM sellout_facts WHERE retailer_id='kruidvat'")}
    assert "999999" not in skus, "het blad 'Sales per day' hoort genegeerd te worden"


def test_bestaande_week_wordt_overgeslagen(omgeving, tmp_path):
    """De kern: een week die al geaggregeerd in de database staat, mag er niet
    per winkel bij komen — de feitensleutel bevat winkel_id, dus beide zouden
    blijven staan en het dashboard zou ze optellen."""
    client, script = omgeving
    import seed
    upload(client, "kv_be.xlsx", seed.make_dwh_xlsx(
        [{"sku": "6369354", "gtin": "4049469072773", "desc": "Slant",
          "brand": "ALESSANDRO", "weeks": {"202540": (10, 100.0)}}],
        country3="BEL", formula="KV"))
    voor = _feiten(client)
    assert voor[1] == 100.0

    weken = ["202540", "202541"]
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(weken), weken))
    assert script.main([pad, "--schrijf"]) == 0

    import db
    with db.get_conn() as conn:
        w40 = conn.execute(
            "SELECT ROUND(SUM(omzet),2) FROM sellout_facts WHERE retailer_id='kruidvat' "
            "AND periode='2025-W40' AND merk='ALESSANDRO'").fetchone()[0]
        w41 = conn.execute(
            "SELECT ROUND(SUM(omzet),2) FROM sellout_facts WHERE retailer_id='kruidvat' "
            "AND periode='2025-W41' AND merk='ALESSANDRO'").fetchone()[0]
    assert w40 == 100.0, "week 40 stond er al geaggregeerd in en mag niet verdubbelen"
    assert w41 == 50.0, "week 41 is nieuw en hoort er per winkel bij te komen"


def test_twee_keer_draaien_verdubbelt_niet(omgeving, tmp_path):
    client, script = omgeving
    weken = ["202540", "202541"]
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(weken), weken))
    script.main([pad, "--schrijf"])
    eerste = _feiten(client)
    script.main([pad, "--schrijf"])
    assert _feiten(client) == eerste

    import db
    with db.get_conn() as conn:
        assert conn.execute("SELECT COUNT(*) FROM imports").fetchone()[0] == 1


def test_analyses_veranderen_niet_van_rekenwijze(omgeving, tmp_path):
    """Kruidvat houdt capability 'winkel' op False, dus de winkelanalyse blijft
    uit en de omzet per winkel blijft op de ingestelde aantallen rekenen. De
    winkeldetails liggen klaar voor later zonder nu ergens door te werken."""
    client, script = omgeving
    import seed
    upload(client, "kv_be.xlsx", seed.make_dwh_xlsx(
        [{"sku": "6369354", "gtin": "4049469072773", "desc": "Slant",
          "brand": "ALESSANDRO", "weeks": {f"2026{w:02d}": (10, 100.0) for w in range(1, 6)}}],
        country3="BEL", formula="KV"))
    weken = ["202540", "202541"]
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(weken), weken))
    script.main([pad, "--schrijf"])

    d = client.get("/api/kruidvat/dashboard").json()
    assert d["capabilities"]["winkel"] is False
    assert d["winkelanalyse"] == {"beschikbaar": False}
    # De historie is er wel bij gekomen.
    assert d["tijdlijn"]["periodes"][0] == "2025-W40"


def test_verkeerd_bestand_wordt_geweigerd(omgeving, tmp_path):
    client, script = omgeving
    import seed
    pad = _bestand(tmp_path, seed.make_dwh_xlsx(
        [{"sku": "1", "gtin": "2", "desc": "X", "brand": "ALESSANDRO",
          "weeks": {"202601": (1, 1.0)}}]))
    with pytest.raises(script.Afgebroken):
        script.lees_bestand(pad)
    assert _feiten(client) == (0, 0)


def test_verwijderen_maakt_de_lading_ongedaan(omgeving, tmp_path):
    """Winkelrijen maken de analyses merkbaar trager; terugdraaien moet dus
    net zo eenvoudig zijn als inlezen — en alleen de eigen lading raken."""
    client, script = omgeving
    import seed
    upload(client, "kv_nl.xlsx", seed.make_dwh_xlsx(
        [{"sku": "31210001", "gtin": "4049469072773", "desc": "Slant",
          "brand": "TWEEZERMAN", "weeks": {"202632": (10, 100.0)}}]))
    ander = _feiten(client)

    weken = ["202540", "202541"]
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(weken), weken))
    script.main([pad, "--schrijf"])
    assert _feiten(client)[0] > ander[0]

    assert script.main([pad, "--verwijder"]) == 0
    # Terug bij af, en de import van een ander bestand staat er nog.
    assert _feiten(client) == ander
    import db
    with db.get_conn() as conn:
        assert conn.execute(
            "SELECT COUNT(*) FROM sellout_facts WHERE winkel_id IS NOT NULL").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM imports").fetchone()[0] == 1


def test_verwijderen_zonder_eerdere_lading_is_ongevaarlijk(omgeving, tmp_path):
    client, script = omgeving
    pad = _bestand(tmp_path, maak_winkelbestand(_twee_winkels(["202540"]), ["202540"]))
    assert script.main([pad, "--verwijder"]) == 0
    assert _feiten(client) == (0, 0)
