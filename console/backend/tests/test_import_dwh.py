"""The import path that actually matters in production: real Kruidvat
DWH-export files through the builtin parser."""

import importlib
import sys
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from test_parser_flow import upload  # noqa: E402

REAL_SAMPLE = Path(
    "/root/.claude/uploads/54377bab-ac94-5cbf-8750-c3a4d90899e0/"
    "aa516215-DWH__Sales_volume__sales_value_per_week_per_article_"
    "Alessendro_Depend_KVNL_5696_1175350483788269736.xlsx")


@pytest.fixture()
def client(tmp_path, monkeypatch):
    monkeypatch.setenv("CONSOLE_DB", str(tmp_path / "console.db"))
    monkeypatch.setenv("CONSOLE_AUTH", "gateway")
    monkeypatch.setenv("CONSOLE_BIND", "127.0.0.1")
    for name in ("db", "seed", "main"):
        sys.modules.pop(name, None)
    main = importlib.import_module("main")
    return TestClient(main.app)


def test_generated_dwh_file_imports(client):
    import seed
    content = seed.make_dwh_xlsx(seed._kv_demo_rows(["202631", "202632"]))
    r = upload(client, "DWH__Sales_volume__sales_Tweezerman_KVNL_32_demo.xlsx", content)
    assert r["status"] == "ingelezen"
    assert r["rows"] == 6           # 3 artikelen x 2 weken
    assert r["retailer_id"] == "kruidvat"

    dash = client.get("/api/kruidvat/dashboard").json()
    assert dash["laatste_periode"] == "2026-W32"
    merken = {b["merk"] for b in dash["kpi"]["omzet"]["breakdown"]}
    assert merken == {"TWEEZERMAN", "ALESSANDRO"}

    art = client.get("/api/kruidvat/artikelen").json()
    assert art["available"] and len(art["artikelen"]) == 3
    eans = {a["ean"] for a in art["artikelen"]}
    assert "31210001" in eans  # GTIN uit het bestand, niet het SKU-nummer

    status = client.get("/api/import-status?retailer_id=kruidvat").json()[0]
    assert status["feeds"], "import status moet de feed tonen"


def test_old_flat_handoff_format_is_no_longer_recognised(client):
    """Het fictieve platte formaat uit het ontwerppakket hoort nu op
    PROFIEL NODIG te stranden: het actieve kruidvat-profiel is de echte
    DWH-parser."""
    from test_parser_flow import make_xlsx
    flat = make_xlsx(["Weeknummer", "EAN", "Merk", "Aantal", "Omzet excl BTW"],
                     [["202632", "31210001", "TWEEZERMAN", 10, 130.0]],
                     sheet="Sellout", meta_rows=8)
    r = upload(client, "DWH_sellout_TWEEZERMAN_NL_wk32.xlsx", flat)
    assert r["status"] in ("profiel_nodig", "error")


def test_unknown_upload_stays_visible_on_retailer_tab(client):
    r = upload(client, "onbekend_rapport_zonder_match.xlsx", b"geen xlsx")
    assert r["status"] == "profiel_nodig"
    rows = client.get("/api/imports?retailer_id=kruidvat").json()
    assert any(x["status"] == "profiel_nodig" for x in rows), \
        "PROFIEL NODIG moet op de retailer-tab zichtbaar zijn"


@pytest.mark.skipif(not REAL_SAMPLE.exists(), reason="echt sample-bestand niet aanwezig")
def test_real_dwh_sample_full_counts(client):
    r = upload(client, REAL_SAMPLE.name[9:], REAL_SAMPLE.read_bytes())
    assert r["status"] == "ingelezen"
    assert r["rows"] == 4566
    dash = client.get("/api/kruidvat/dashboard").json()
    per_brand = {b["merk"]: b["waarde"] for b in dash["kpi"]["omzet"]["breakdown"]}
    # Het bestand schrijft "DEPEND GEL IQ"; de app voegt dat samen met DEPEND
    # zoals ICI het levert (engine/merken.py), anders zijn het twee merken.
    assert set(per_brand) == {"ALESSANDRO", "DEPEND"}
    art = client.get("/api/kruidvat/artikelen").json()
    assert len(art["artikelen"]) == 156


@pytest.mark.skipif(not REAL_SAMPLE.exists(), reason="echt sample-bestand niet aanwezig")
def test_real_file_with_mangled_name_still_recognised(client):
    """Mailclients en downloads plakken voorvoegsels aan bestandsnamen; de
    inhoudsherkenning moet het bestand dan alsnog bij kruidvat brengen."""
    r = upload(client, "aa516215-hernoemd en (1) gek gemaakt.xlsx", REAL_SAMPLE.read_bytes())
    assert r["status"] == "ingelezen" and r["retailer_id"] == "kruidvat"


def test_dwh_total_mismatch_imports_with_visible_warning(client):
    """Gelijk aan de Streamlit-app: een niet-kloppend Total-getal blokkeert de
    import niet, maar levert wel een waarschuwing die de UI toont — blokkeren
    zou betekenen dat je bij een bronbestandsfout helemaal geen cijfers hebt."""
    import io
    import seed
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(seed.make_dwh_xlsx(seed._kv_demo_rows(["202632"]))))
    ws = wb.active
    ws.cell(row=ws.max_row, column=5, value=999999)     # Total-rij verminken
    out = io.BytesIO()
    wb.save(out)

    r = upload(client, "DWH__Sales_volume_TWEEZERMAN_KVNL_kapot-totaal.xlsx", out.getvalue())
    assert r["status"] == "ingelezen"
    assert "Total-rij" in r["detail"]          # waarschuwing, zichtbaar in de UI
    assert client.get("/api/kruidvat/dashboard").json()["empty"] is False


def test_dwh_wrong_sheet_choice_still_fails_closed(client):
    """De ene uitzondering: sluit GEEN enkel blad aan op zijn Total-rij, dan
    is niet vast te stellen welk blad gezaghebbend is en zou het verkeerde
    blad elke SKU per GTIN-variant dubbel tellen."""
    import io
    import seed
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(seed.make_dwh_xlsx(seed._kv_demo_rows(["202632"]))))
    ws = wb.active
    ws.cell(row=ws.max_row, column=3, value=99)   # SKU-telling in Total-rij verminken
    out = io.BytesIO()
    wb.save(out)

    r = upload(client, "DWH__Sales_volume_TWEEZERMAN_KVNL_verkeerd-blad.xlsx", out.getvalue())
    assert r["status"] == "error" and "SKU-telling" in r["detail"]


def test_history_is_preserved_and_analysis_updates_per_import(client):
    """Het werkmodel: per retailer een vaste parser; elke import werkt de
    analyse bij, vult de historie aan en overschrijft alleen de periodes die
    in het nieuwe bestand zitten."""
    import seed

    def kv(week_values: dict) -> bytes:
        """week_values: {"202631": (volume, omzet)} voor één artikel."""
        return seed.make_dwh_xlsx([{
            "sku": "31210001", "gtin": "4049469072773",
            "desc": "Tweezerman Slant Tweezer", "brand": "TWEEZERMAN",
            "weeks": week_values}])

    # 1. Eerste levering: weken 30 en 31.
    assert upload(client, "DWH__Sales_Tweezerman_KVNL_wk31.xlsx",
                  kv({"202630": (10, 130.0), "202631": (12, 156.0)}))["status"] == "ingelezen"
    dash = client.get("/api/kruidvat/dashboard").json()
    assert dash["laatste_periode"] == "2026-W31"
    assert dash["kpi"]["omzet"]["waarde"] == pytest.approx(156.0)

    # 2. Volgende levering: week 32. Historie blijft, analyse schuift op.
    assert upload(client, "DWH__Sales_Tweezerman_KVNL_wk32.xlsx",
                  kv({"202632": (14, 182.0)}))["status"] == "ingelezen"
    dash = client.get("/api/kruidvat/dashboard").json()
    assert dash["laatste_periode"] == "2026-W32"
    serie = dash["trend"]["series"]["omzet"]["2026"]
    assert {"30", "31", "32"} <= set(serie), "eerdere weken moeten bewaard blijven"
    assert serie["30"] == pytest.approx(130.0)

    # 3. Correctie op week 31: alleen die week verandert.
    assert upload(client, "DWH__Sales_Tweezerman_KVNL_wk31-correctie.xlsx",
                  kv({"202631": (12, 999.0)}))["status"] == "ingelezen"
    serie = client.get("/api/kruidvat/dashboard").json()["trend"]["series"]["omzet"]["2026"]
    assert serie["31"] == pytest.approx(999.0), "week 31 moet bijgewerkt zijn"
    assert serie["30"] == pytest.approx(130.0), "week 30 mag niet veranderen"
    assert serie["32"] == pytest.approx(182.0), "week 32 mag niet veranderen"

    # 4. Artikelanalyse volgt dezelfde dataset.
    art = client.get("/api/kruidvat/artikelen").json()
    assert art["available"] and len(art["artikelen"]) == 1
    assert art["artikelen"][0]["totaal_ytd"]["omzet"] == pytest.approx(130.0 + 999.0 + 182.0)


def test_correction_for_one_brand_leaves_other_brands_untouched(client):
    """De gevaarlijkste variant: een losse levering per merk voor dezelfde
    week mag het andere merk niet wissen én niet dubbel tellen."""
    import seed

    def kv(brand, gtin, sku, week, vol, omzet):
        return seed.make_dwh_xlsx([{
            "sku": sku, "gtin": gtin, "desc": f"{brand} artikel",
            "brand": brand, "weeks": {week: (vol, omzet)}}])

    assert upload(client, "DWH__Sales_Tweezerman_KVNL_a.xlsx",
                  kv("TWEEZERMAN", "31210001", "31210001", "202632", 10, 100.0))["status"] == "ingelezen"
    assert upload(client, "DWH__Sales_Alessandro_KVNL_b.xlsx",
                  kv("ALESSANDRO", "31210003", "31210003", "202632", 5, 50.0))["status"] == "ingelezen"

    per_merk = {b["merk"]: b["waarde"] for b in
                client.get("/api/kruidvat/dashboard").json()["kpi"]["omzet"]["breakdown"]}
    assert per_merk == pytest.approx({"TWEEZERMAN": 100.0, "ALESSANDRO": 50.0})

    # Correctie op alleen TWEEZERMAN: ALESSANDRO blijft ongewijzigd staan.
    assert upload(client, "DWH__Sales_Tweezerman_KVNL_a-correctie.xlsx",
                  kv("TWEEZERMAN", "31210001", "31210001", "202632", 10, 300.0))["status"] == "ingelezen"
    per_merk = {b["merk"]: b["waarde"] for b in
                client.get("/api/kruidvat/dashboard").json()["kpi"]["omzet"]["breakdown"]}
    assert per_merk == pytest.approx({"TWEEZERMAN": 300.0, "ALESSANDRO": 50.0})
