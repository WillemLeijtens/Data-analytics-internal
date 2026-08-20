"""Ingebouwde parser voor de echte Etos "Data Grid"-widgetexport
("Sales per item analyse").

Het bestand is één tabblad met een metadatablok bovenaan en daaronder een
artikel×week-matrix:

  * metadata: "Brand (N)" plus een Time-scope die per export verschilt —
    "Fiscal YTD 202601-202632", 'Fiscal Quarter "202504 (Weeks
    202541-202552, Ending 28/12/2025)"', '2 Fiscal Quarters 202502-202503,
    Ending 05/10/2025' of '9 Fiscal Periods 202405-202413, Ending
    29/12/2024'. Dit is de enige controle-informatie die het formaat biedt
    (er is géén totalenrij zoals bij Kruidvat of merk-tab zoals bij ICI),
    dus de parser verifieert per scope alles wat eruit af te leiden valt:
    een expliciete week-range wordt exact afgedwongen, een Ending-datum
    moet de ISO-zondag van de laatste weekkolom zijn, en de weekreeks moet
    altijd aaneengesloten zijn.
  * weekkoppen: "202601 (Ending 04/01/2026)" — elke week 2 kolommen breed.
    De Ending-datums vallen exact op de ISO-zondag van die week; dat wordt
    per week gecontroleerd, zodat een stille overstap van Etos op een
    andere weekkalender de import laat FALEN in plaats van weken verkeerd
    te labelen.
  * subkoppen: UPC Name | UPC ID | Brand | [Sales € TY | Units TY] × weken.

Feiten: omzet + volume per artikel per week per merk. Land is constant NL
(Etos is alleen in Nederland actief); winkel- en bannerniveau levert dit
formaat niet. Merk komt binnen als "TWEEZERMAN - 2278"; het interne nummer
wordt gestript zodat merknamen matchen met de andere retailers.
"""

from __future__ import annotations

import datetime as dt
import io
import math
import re

from openpyxl import load_workbook

from .cellen import als_identifier
from .periods import parse_period

WEEK_HDR_RE = re.compile(r"^(\d{6})\s*\(Ending\s+(\d{2}/\d{2}/\d{4})\)$")
# Expliciete week-ranges, in volgorde van voorkeur: de Quarter-variant noemt
# de weken letterlijk ("Weeks 202541-202552"), de YTD-variant als
# "Fiscal YTD 202601-202632". Quarters/Periods-selecties zonder week-range
# vallen terug op de intrinsieke aaneengesloten-reeks-eis.
WEEKS_RANGE_RE = re.compile(r"Weeks\s+(\d{6})-(\d{6})")
YTD_RANGE_RE = re.compile(r"Fiscal\s+YTD\s+(\d{6})-(\d{6})")
ENDING_RE = re.compile(r"Ending\s+(\d{2}/\d{2}/\d{4})")
BRAND_COUNT_RE = re.compile(r"Brand\s*\((\d+)\)")
BRAND_SUFFIX_RE = re.compile(r"\s*-\s*\d+$")


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _num(raw):
    """Getal met punt-decimaal (zoals deze export levert); leeg -> None."""
    if raw is None or raw == "" or isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        number = float(raw)
        return number if math.isfinite(number) else None
    s = str(raw).strip().replace("€", "").replace(" ", "").replace(" ", "")
    if not s:
        return None
    # Deze export levert punt-decimaal ("1,234.56": komma is duizendtal),
    # maar een cel in Europese notatie ("1.234,56") zou met alleen komma's
    # strippen stil een factor duizend misgaan. Staan beide tekens erin,
    # dan is de laatste het decimaalteken — zelfde regel als de
    # Kruidvat-parser hanteert.
    if "," in s and "." in s:
        decimaal = "," if s.rfind(",") > s.rfind(".") else "."
        duizend = "." if decimaal == "," else ","
        s = s.replace(duizend, "").replace(decimaal, ".")
    else:
        s = s.replace(",", "")
    number = float(s)
    return number if math.isfinite(number) else None


def _rows(content: bytes) -> list[list]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [["" if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]


def _find_subheader(rows) -> int | None:
    for i, row in enumerate(rows[:40]):
        vals = {_norm(c) for c in row}
        if {"UPC Name", "UPC ID", "Brand"} <= vals:
            return i
    return None


def content_matches(content: bytes) -> bool:
    """Structuurherkenning, zodat een hernoemd bestand herkend blijft: de
    UPC-subkopregel mét weekkoppen ("… (Ending …)") in de rij erboven."""
    try:
        rows = _rows(content)
    except Exception:  # noqa: BLE001 - onleesbaar bestand matcht simpelweg niet
        return False
    i = _find_subheader(rows)
    if i is None or i == 0:
        return False
    return any(WEEK_HDR_RE.match(_norm(c)) for c in rows[i - 1])


def _expected_weeks(start: str, end: str) -> list[str]:
    """Alle yyyyww-waarden van start t/m end, over een jaargrens heen als
    het fiscale jaar ooit door de kalenderjaargrens loopt."""
    y, w = int(start[:4]), int(start[4:])
    ey, ew = int(end[:4]), int(end[4:])
    out = []
    while (y, w) <= (ey, ew):
        out.append(f"{y}{w:02d}")
        try:
            dt.date.fromisocalendar(y, w + 1, 1)
            w += 1
        except ValueError:
            y, w = y + 1, 1
        if len(out) > 120:
            raise ValueError("weekbereik in de metadata is onaannemelijk groot")
    return out


def parse_workbook(content: bytes) -> dict:
    rows = _rows(content)
    sub_i = _find_subheader(rows)
    if sub_i is None or sub_i == 0:
        raise ValueError("geen UPC-kopregel gevonden — is dit wel een "
                         "Etos Data Grid-export?")
    subhdr = [_norm(c) for c in rows[sub_i]]
    weekhdr = [_norm(c) for c in rows[sub_i - 1]]

    # Metadata: de enige controlebron van dit formaat. Het merkental is in
    # elke scope-variant aanwezig en dus verplicht; de week-range hangt af
    # van de gekozen Time-selectie.
    # NB: de weekkoppenrij zelf (rij boven de subkop) telt niet mee als
    # metadata — anders zou een Ending-datum daaruit zijn eigen controle
    # kunnen "bevestigen".
    blok = " | ".join(_norm(c) for r in rows[:sub_i - 1] for c in r if _norm(c))
    m_brands = BRAND_COUNT_RE.search(blok)
    if not m_brands:
        raise ValueError("metadatablok mist 'Brand (N)' — zonder dat "
                         "controlegetal is het merkental niet te verifiëren; "
                         "import afgebroken")
    verwacht_merken = int(m_brands.group(1))
    m_range = WEEKS_RANGE_RE.search(blok) or YTD_RANGE_RE.search(blok)
    scope_endings = set(ENDING_RE.findall(blok))

    # Weekkolommen + de drie ankers uit de subkop.
    try:
        col_naam = subhdr.index("UPC Name")
        col_upc = subhdr.index("UPC ID")
        col_merk = subhdr.index("Brand")
    except ValueError as e:
        raise ValueError(f"subkop mist een verplichte kolom: {e}")
    weekcols: list[tuple[str, int]] = []      # (canonieke periode, sales-kolom)
    gezien: list[str] = []
    for j, cel in enumerate(weekhdr):
        m = WEEK_HDR_RE.match(cel)
        if not m:
            continue
        raw, ending = m.group(1), m.group(2)
        periode = parse_period(raw, "yyyyww")   # incl. week-53-validatie
        jaar, week = int(raw[:4]), int(raw[4:])
        # Ending-datum moet de ISO-zondag zijn: wijkt Etos af van ISO-weken,
        # dan is elk weeklabel verdacht en stoppen we.
        eind = dt.datetime.strptime(ending, "%d/%m/%Y").date()
        if eind != dt.date.fromisocalendar(jaar, week, 7):
            raise ValueError(
                f"weekkop {raw}: Ending {ending} is niet de ISO-zondag van "
                f"die week — weekkalender wijkt af; import afgebroken")
        if subhdr[j] != "Sales € TY" or (j + 1 >= len(subhdr)) or subhdr[j + 1] != "Units TY":
            raise ValueError(f"weekkop {raw}: verwachtte 'Sales € TY'+'Units TY' "
                             "eronder; kolomindeling wijkt af")
        gezien.append(raw)
        weekcols.append((periode, j))
    if not gezien:
        raise ValueError("geen weekkolommen gevonden boven de UPC-kopregel")
    if m_range:
        # Expliciete range in de metadata ("Weeks …-…" of "Fiscal YTD …-…"):
        # de kolommen moeten daar exact op aansluiten.
        verwachte_weken = _expected_weeks(m_range.group(1), m_range.group(2))
        bron = f"metadata-range {m_range.group(1)}-{m_range.group(2)}"
    else:
        # Quarters/Periods zonder week-range: dan geldt de intrinsieke eis
        # dat de reeks aaneengesloten en oplopend is — een gat betekent een
        # kapotte of geknipte export.
        verwachte_weken = _expected_weeks(gezien[0], gezien[-1])
        bron = f"aaneengesloten reeks {gezien[0]}-{gezien[-1]}"
    if gezien != verwachte_weken:
        ontbreekt = sorted(set(verwachte_weken) - set(gezien))
        extra = sorted(set(gezien) - set(verwachte_weken))
        raise ValueError(
            f"weekkolommen sluiten niet aan op de {bron}"
            + (f"; ontbreekt: {', '.join(ontbreekt)}" if ontbreekt else "")
            + (f"; onverwacht: {', '.join(extra)}" if extra else ""))
    if scope_endings:
        # De Time-scope noemt een einddatum; die moet de ISO-zondag van de
        # laatste weekkolom zijn — anders gaan metadata en kolommen niet
        # over dezelfde periode.
        laatste_raw = gezien[-1]
        laatste_zondag = dt.date.fromisocalendar(
            int(laatste_raw[:4]), int(laatste_raw[4:]), 7).strftime("%d/%m/%Y")
        if laatste_zondag not in scope_endings:
            raise ValueError(
                f"de einddatum in de Time-scope ({', '.join(sorted(scope_endings))}) "
                f"hoort niet bij de laatste weekkolom {laatste_raw} "
                f"(zondag {laatste_zondag}); metadata en kolommen gaan niet "
                "over dezelfde periode — import afgebroken")

    facts: list[dict] = []
    warnings: list[str] = []
    # Een lege Sales- of Units-cel werd stilzwijgend als 0 geboekt: "niets
    # verkocht" is dan niet te onderscheiden van "niet geleverd", en dat
    # verschil stuurt wél de volume-KPI, de rotatie en de prijsindex aan.
    # Zelfde melding als de Kruidvat-parser al gaf.
    lege_cellen = 0
    for r in rows[sub_i + 1:]:
        upc_raw = r[col_upc] if col_upc < len(r) else ""
        merk_raw = _norm(r[col_merk] if col_merk < len(r) else "")
        # Disclaimer- en lege regels onder de tabel hebben geen UPC+merk. De
        # numerieke toets blijft staan: hij bepaalt of dit überhaupt een
        # datarij is. De WAARDE komt daarna uit als_identifier(), want
        # str(int(...)) maakte van UPC "012345678905" stil "12345678905".
        if _num(upc_raw) is None or not merk_raw:
            continue
        upc = als_identifier(upc_raw)
        if upc is None:
            continue
        merk = BRAND_SUFFIX_RE.sub("", merk_raw).strip().upper()
        naam = _norm(r[col_naam] if col_naam < len(r) else "") or None
        for periode, j in weekcols:
            sales = _num(r[j] if j < len(r) else None)
            units = _num(r[j + 1] if j + 1 < len(r) else None)
            if sales is None and units is None:
                continue                       # geen verkoop die week
            if sales is None or units is None:
                lege_cellen += 1
            facts.append({
                "periode": periode, "land": "NL", "banner": None,
                "winkel_id": None, "winkel_naam": None,
                "merk": merk, "artikel_ean": upc, "artikel_naam": naam,
                "volume": int(round(units or 0)), "omzet": float(sales or 0.0),
            })

    if not facts:
        raise ValueError("geen artikelrijen met cijfers gevonden — is dit wel "
                         "een gevulde Data Grid-export?")

    # Consistentie, fail-closed — het maximum dat dit formaat te bieden heeft.
    keys = [(f["artikel_ean"], f["periode"]) for f in facts]
    if len(keys) != len(set(keys)):
        dubbel = sorted({k for k in keys if keys.count(k) > 1})[:5]
        raise ValueError(f"dubbele artikel/week-combinaties in het bestand "
                         f"(o.a. {dubbel}); import afgebroken om dubbeltelling "
                         "te voorkomen")
    merken = {f["merk"] for f in facts}
    if len(merken) != verwacht_merken:
        raise ValueError(
            f"metadata zegt Brand ({verwacht_merken}) maar de tabel bevat "
            f"{len(merken)} merk(en): {', '.join(sorted(merken))}; "
            "import afgebroken")

    if lege_cellen:
        warnings.append(
            f"{lege_cellen} Sales/Units-cel(len) waren leeg terwijl de andere "
            "helft van het paar wél gevuld was, en zijn als 0 geboekt — "
            "controleer het bronbestand voordat je hierop stuurt.")

    return {"facts": facts, "periode_type": "week",
            "periodes": sorted({f["periode"] for f in facts}),
            "warnings": warnings}
