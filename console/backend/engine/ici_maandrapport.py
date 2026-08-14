"""Ingebouwde parser voor het echte ICI Paris XL-maandrapport.

Het bestand is een rapport met meerdere tabbladen; de parser kiest de juiste
tabs op STRUCTUUR, niet op naam of positie — dus ook als ICI tabs hernoemt
of verschuift blijft de herkenning werken:

  * winkel-tabs: elk merkblok heeft een kopregel met "Store" + "Address"
    gevolgd door maandkolommen (YYYYMM); de merknaam staat er enkele rijen
    boven. Meerdere blokken onder elkaar = meerdere merken.
  * merk-tab ("Brands"): kopregel "Year" + "Category" + maandnummers 01-12 —
    gebruikt om de winkelcijfers per merk/maand te RECONCILIËREN, zoals de
    Kruidvat-parser dat met de Total-rij doet.

Feiten: omzet per winkel per maand per merk. Volume levert ICI niet
(volume=0, capability-vlag volume=False stuurt de UI). "Total"-kolommen en
groeipercentages worden overgeslagen; alleen YYYYMM-kolommen tellen.
"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook

from .periods import PeriodError, parse_period

MONTH_RE = re.compile(r"^\d{6}$")


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _is_store_header(row) -> bool:
    vals = {_norm(c) for c in row}
    return "Store" in vals and "Address" in vals


def _month_columns(row) -> list[tuple[int, str]]:
    out = []
    for j, c in enumerate(row):
        s = _norm(c)
        if MONTH_RE.fullmatch(s):
            try:
                out.append((j, parse_period(s, "yyyymm")))
            except PeriodError:
                continue
    return out


def _find_brand_above(rows, header_idx: int) -> str | None:
    """De merknaam staat 1-4 rijen boven de blok-kopregel, als enige tekst."""
    for i in range(header_idx - 1, max(-1, header_idx - 5), -1):
        for c in rows[i][:6]:
            s = _norm(c)
            if s and not MONTH_RE.fullmatch(s) and not s.isdigit() \
                    and s not in ("Store", "Address", "Total"):
                return s.upper()
    return None


def _to_number(raw):
    if raw is None or raw == "":
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    try:
        return float(str(raw).strip().replace(",", "."))
    except ValueError:
        return None


def content_matches(content: bytes) -> bool:
    """Herkenning op inhoud: een tabblad met een Store/Address-kopregel
    gevolgd door minstens één YYYYMM-maandkolom."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=1, max_row=12, values_only=True):
                if _is_store_header(row) and _month_columns(row):
                    return True
        return False
    except Exception:  # noqa: BLE001 - onleesbaar bestand matcht simpelweg niet
        return False


def parse_workbook(content: bytes) -> dict:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)

    facts: list[dict] = []
    warnings: list[str] = []
    brand_totals: dict[tuple[str, str], float] = {}   # (merk, periode) -> rapporttotaal

    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        # Merk-tab: Year | Category | 01..12 — voor reconciliatie.
        for i, row in enumerate(rows[:6]):
            vals = [_norm(c) for c in row]
            if "Year" in vals and "Category" in vals:
                y_col = vals.index("Year")
                b_col = vals.index("Category")
                mcols = [(j, int(v)) for j, v in enumerate(vals)
                         if v.isdigit() and len(v) == 2 and 1 <= int(v) <= 12]
                year = None
                for r in rows[i + 1:]:
                    year = int(r[y_col]) if _norm(r[y_col]).isdigit() else year
                    brand = _norm(r[b_col]).upper()
                    if not brand or year is None:
                        continue
                    for j, m in mcols:
                        v = _to_number(r[j] if j < len(r) else None)
                        if v is not None:
                            brand_totals[(brand, f"{year}-{m:02d}")] = v
                break

        # Winkel-tabs: blokken met Store/Address-kop + maandkolommen.
        for i, row in enumerate(rows):
            if not _is_store_header(row):
                continue
            months = _month_columns(row)
            if not months:
                continue
            brand = _find_brand_above(rows, i)
            if not brand:
                warnings.append(f"Blad {ws.title!r}: winkelblok op rij {i + 1} "
                                "zonder herkenbare merknaam — overgeslagen.")
                continue
            store_col = [_norm(c) for c in row].index("Store")
            name_col = [_norm(c) for c in row].index("Address")
            for r in rows[i + 1:]:
                if _is_store_header(r):        # volgend blok
                    break
                store = _norm(r[store_col] if store_col < len(r) else None)
                if not store:
                    continue
                if not store.replace(".", "").isdigit():
                    break                       # einde blok (bijv. totaalregel)
                store = str(int(float(store)))
                naam = _norm(r[name_col] if name_col < len(r) else None) or None
                for j, periode in months:
                    v = _to_number(r[j] if j < len(r) else None)
                    if v is None:
                        continue
                    facts.append({
                        "periode": periode, "land": "NL", "banner": None,
                        "winkel_id": store, "winkel_naam": naam,
                        "merk": brand, "artikel_ean": None, "artikel_naam": None,
                        "volume": 0, "omzet": v,
                    })

    if not facts:
        raise ValueError("Geen winkelblokken met maandkolommen gevonden — is dit "
                         "wel een ICI Paris XL-maandrapport?")

    # Reconciliatie tegen de merk-tab, zoals Kruidvat tegen de Total-rij.
    if brand_totals:
        computed: dict[tuple[str, str], float] = {}
        for f in facts:
            k = (f["merk"], f["periode"])
            computed[k] = computed.get(k, 0.0) + f["omzet"]
        mismatches = sum(
            1 for k, expected in brand_totals.items()
            if k in computed and abs(computed[k] - expected) > max(1.0, 0.005 * abs(expected)))
        if mismatches:
            warnings.append(
                f"{mismatches} merk/maand-combinatie(s) sluiten niet aan op de "
                "totalen in de merk-tab — controleer het bestand.")

    return {"facts": facts, "periode_type": "maand",
            "periodes": sorted({f["periode"] for f in facts}),
            "warnings": warnings}
