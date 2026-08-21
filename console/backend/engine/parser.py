"""File detection + parsing against a parser profile (PROMPT.md §3).

Detection: filename_glob first, sheet/required_headers as tiebreak. Zero or
multiple matches -> the import gets status 'profiel_nodig'. Only live/test
profiles take part in detection; concept profiles exist to be completed in
the Parser screen.

Parsing is transactional at the caller: this module either returns the full
list of canonical fact dicts or raises ParseError with row-level detail —
never a partial result.
"""

from __future__ import annotations

import csv
import fnmatch
import io
import math
import re

from .periods import PeriodError, parse_period
from .profile import Profile, capabilities

# EAN-8, UPC-A (12) en EAN-13. UPC-A stond er eerst niet bij, waardoor een
# retailer die 12-cijferige codes levert een volledige importweigering zou
# krijgen — terwijl dat een volstrekt normale artikelcode is.
EAN_RE = re.compile(r"\d{8}$|\d{12}$|\d{13}$")


class ParseError(Exception):
    def __init__(self, message: str, row_errors: list[dict] | None = None):
        super().__init__(message)
        self.row_errors = row_errors or []


# ---------------------------------------------------------------- reading

def _read_csv(content: bytes, delimiter: str) -> list[list[str]]:
    text = content.decode("utf-8-sig")
    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter or ";")]


def _read_xlsx(content: bytes, sheet: str | None) -> list[list]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    if sheet and sheet in wb.sheetnames:
        ws = wb[sheet]
    elif sheet:
        raise ParseError(f"werkblad {sheet!r} niet gevonden (wel: {', '.join(wb.sheetnames)})")
    else:
        ws = wb[wb.sheetnames[0]]
    return [["" if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]


def read_table(content: bytes, detection: dict) -> tuple[list[str], list[list]]:
    """Returns (headers, data_rows) using the profile's header_row (1-based)."""
    if detection.get("filetype") == "csv":
        rows = _read_csv(content, detection.get("csv_delimiter"))
    else:
        rows = _read_xlsx(content, detection.get("sheet"))
    header_row = int(detection.get("header_row") or 1)
    if len(rows) < header_row:
        raise ParseError(f"bestand heeft geen rij {header_row} (header-rij volgens profiel)")
    headers = [str(h).strip() for h in rows[header_row - 1]]
    data = [r for r in rows[header_row:] if any(str(c).strip() for c in r)]
    return headers, data


def probe_headers(content: bytes, detection: dict) -> list[str] | None:
    try:
        headers, _ = read_table(content, detection)
        return headers
    except Exception:  # noqa: BLE001 - a failed probe just means "no match"
        return None


# ---------------------------------------------------------------- detection

def sniff(filename: str, content: bytes) -> dict | None:
    """Best-effort look inside an UNRECOGNISED file, so the Parser screen can
    offer its real columns instead of an empty mapping table.

    Returns {filetype, sheet, header_row, csv_delimiter, columns, voorbeeld}
    or None when nothing table-like is found. Never raises: a file we cannot
    read simply yields no suggestion.
    """
    name = filename.lower()
    try:
        if name.endswith((".csv", ".txt", ".tsv")):
            return _sniff_csv(content)
        if name.endswith((".xlsx", ".xlsm")):
            return _sniff_xlsx(content)
        # Unknown extension: try both, spreadsheet first.
        return _sniff_xlsx(content) or _sniff_csv(content)
    except Exception:  # noqa: BLE001 - a failed sniff is not an error
        return None


def _score(cells: list) -> int:
    """How much a row looks like a header: non-empty, non-numeric labels."""
    score = 0
    for c in cells:
        s = str(c).strip()
        if not s or s.startswith("#"):
            continue
        try:
            float(s.replace(",", "."))
        except ValueError:
            score += 1
    return score


def _pick_header(rows: list[list], limit: int = 25) -> tuple[int, list[str]] | None:
    """Header = the highest-scoring row in the first `limit` rows; ties go to
    the earliest. Handles metadata blocks above the real header."""
    best = None
    for i, row in enumerate(rows[:limit]):
        s = _score(row)
        if s >= 2 and (best is None or s > best[0]):
            best = (s, i)
    if best is None:
        return None
    idx = best[1]
    headers = [str(c).strip() for c in rows[idx]]
    while headers and not headers[-1]:
        headers.pop()
    return idx, headers


def _sniff_csv(content: bytes) -> dict | None:
    text = content.decode("utf-8-sig", errors="replace")
    best = None
    for delim in (";", ",", "\t", "|"):
        rows = [r for r in csv.reader(io.StringIO(text), delimiter=delim)]
        rows = [r for r in rows if not (r and str(r[0]).startswith("#"))]
        picked = _pick_header(rows)
        if picked and (best is None or len(picked[1]) > len(best[1][1])):
            best = (delim, picked, rows)
    if not best or len(best[1][1]) < 2:
        return None
    delim, (idx, headers), rows = best
    sample = rows[idx + 1] if len(rows) > idx + 1 else []
    return {"filetype": "csv", "sheet": None, "header_row": idx + 1,
            "csv_delimiter": delim, "columns": headers,
            "voorbeeld": [str(c) for c in sample[:len(headers)]]}


def _sniff_xlsx(content: bytes) -> dict | None:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    best = None
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            rows.append(["" if c is None else c for c in row])
            if i >= 30:
                break
        picked = _pick_header(rows)
        if picked and (best is None or len(picked[1]) > len(best[1][1])):
            best = (sheet, picked, rows)
    if not best or len(best[1][1]) < 2:
        return None
    sheet, (idx, headers), rows = best
    sample = rows[idx + 1] if len(rows) > idx + 1 else []
    return {"filetype": "xlsx", "sheet": sheet, "header_row": idx + 1,
            "csv_delimiter": None, "columns": headers,
            "voorbeeld": [str(c) for c in sample[:len(headers)]]}


def _effective(profiles: list[Profile]) -> list[Profile]:
    """One candidate per retailer: the newest LIVE version, else the newest
    TEST one. Without this, every re-publish would make a retailer's own
    older version compete with the new one and turn detection ambiguous."""
    best: dict[str, Profile] = {}
    rank = {"live": 2, "test": 1}
    for p in profiles:
        r = rank.get(p.status, 0)
        if r == 0:
            continue
        cur = best.get(p.retailer_id)
        if cur is None or (rank[cur.status], cur.version) < (r, p.version):
            best[p.retailer_id] = p
    return list(best.values())


def kandidaten(filename: str, content: bytes, profiles: list[Profile]) -> list[Profile]:
    """Elk profiel dat dit bestand herkent — één, geen, of meerdere.

    `detect` geeft bij meerdere kandidaten None terug: liever niets doen dan
    gokken. Maar "geen parser herkent dit" en "twee retailers herkennen dit"
    vragen om een heel ander antwoord aan de gebruiker: het tweede geval is op
    te lossen door te kiezen. Twee ICI-rapporten (NL en BE) hebben dezelfde
    structuur en zijn inhoudelijk niet uit elkaar te houden — alleen de
    aanleveraar weet welk land het is.
    """
    profiles = _effective(profiles)
    op_naam = [p for p in profiles
               if fnmatch.fnmatch(filename.lower(),
                                  p.definition["detection"]["filename_glob"].lower())]
    if len(op_naam) > 1:
        bevestigd = [p for p in op_naam if _headers_match(content, p)]
        op_naam = bevestigd or op_naam
    if op_naam:
        return op_naam
    return [p for p in profiles
            if (_builtin_content_match(content, p) if p.definition.get("builtin")
                else _headers_match(content, p))]


def detect(filename: str, content: bytes, profiles: list[Profile]) -> Profile | None:
    """Pick the profile for a file. None => 'profiel_nodig'."""
    profiles = _effective(profiles)
    candidates = [p for p in profiles
                  if fnmatch.fnmatch(filename.lower(),
                                     p.definition["detection"]["filename_glob"].lower())]
    if len(candidates) > 1:
        # Tiebreak on sheet + required headers actually present.
        confirmed = [p for p in candidates if _headers_match(content, p)]
        candidates = confirmed if confirmed else candidates
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        return None  # ambiguous — user must map/choose once
    # No filename match: try content-based match as a last resort. Builtin
    # profiles recognise their format by structure, so a renamed or
    # prefix-mangled file (mail clients, downloads) still lands correctly.
    confirmed = [p for p in profiles
                 if (_builtin_content_match(content, p) if p.definition.get("builtin")
                     else _headers_match(content, p))]
    return confirmed[0] if len(confirmed) == 1 else None


def _builtin_content_match(content: bytes, profile: Profile) -> bool:
    """Structure probes per builtin parser, so renamed files still land at
    the right retailer."""
    builtin = profile.definition.get("builtin")
    if builtin == "ici_maandrapport":
        from . import ici_maandrapport
        return ici_maandrapport.content_matches(content)
    if builtin == "etos_datagrid":
        from . import etos_datagrid
        return etos_datagrid.content_matches(content)
    if builtin != "kruidvat_dwh":
        return False
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for ws in wb.worksheets:
            found_sku = found_formula = found_country = False
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
                for v in row:
                    if v == "SKU No.":
                        found_sku = True
                    elif v == "Formula:" and i < 10:
                        found_formula = True
                    elif v == "Country:" and i < 10:
                        found_country = True
            if found_sku and found_formula and found_country:
                return True
        return False
    except Exception:  # noqa: BLE001 - unreadable file simply doesn't match
        return False


def _headers_match(content: bytes, profile: Profile) -> bool:
    det = profile.definition["detection"]
    req = det.get("required_headers")
    if not req:
        # No required headers = nothing to confirm on. Without this guard an
        # empty list would "match" every readable file (all([]) is True) and
        # a builtin profile would swallow other retailers' uploads.
        return False
    headers = probe_headers(content, det)
    if headers is None:
        return False
    hset = {h.lower() for h in headers}
    return all(r.lower() in hset for r in req)


# ---------------------------------------------------------------- numbers

def parse_number(value, decimal: str) -> float:
    """Parse with the profile's decimal sign; both '1.234,56' and '1234.56'."""
    if isinstance(value, bool):
        raise ValueError("boolean is geen getal")
    if isinstance(value, (int, float)):
        number = float(value)
        if not math.isfinite(number):
            raise ValueError("getal is niet eindig")
        return number
    s = (str(value).strip().replace("€", "").replace(" ", "")
         .replace(" ", "").replace(" ", "").replace("'", ""))
    if not s:
        raise ValueError("leeg")

    # Staan beide scheidingstekens erin, dan is de laatste het decimaalteken.
    # Bij één scheidingsteken volgen we het profiel, tenzij het patroon
    # onmiskenbaar duizendtallen zijn — dat vangt de veelvoorkomende
    # spreadsheet-uitschieter op (punt-decimaal in een komma-export).
    if "," in s and "." in s:
        decimal_seen = "," if s.rfind(",") > s.rfind(".") else "."
        thousands = "." if decimal_seen == "," else ","
        s = s.replace(thousands, "").replace(decimal_seen, ".")
    else:
        separator = "," if "," in s else "." if "." in s else None
        if separator:
            grouped = re.fullmatch(
                rf"[+-]?\d{{1,3}}(?:{re.escape(separator)}\d{{3}})+", s)
            if separator != decimal and grouped:
                s = s.replace(separator, "")
            else:
                s = s.replace(separator, ".")
    number = float(s)
    if not math.isfinite(number):
        raise ValueError("getal is niet eindig")
    return number


# ---------------------------------------------------------------- parsing

def _parse_builtin_kruidvat(filename: str, content: bytes) -> dict:
    """Bridge to the vendored, battle-tested Kruidvat DWH parser: the real
    export (metadata block + side-by-side week columns) cannot be expressed
    as a column mapping."""
    from . import kruidvat_dwh
    from .periods import parse_period
    try:
        parsed = kruidvat_dwh.parse_workbook(io.BytesIO(content), filename)
    except ValueError as e:
        raise ParseError(str(e))
    except Exception as e:  # noqa: BLE001 - onleesbaar bestand is geen crash
        raise ParseError(f"bestand kon niet worden gelezen: {e}") from e
    iteminfo = {(i["brand"], i["sku"]): i for i in parsed.items}
    facts = []
    for f in parsed.facts:
        info = iteminfo.get((f["brand"], f["unit"]), {})
        facts.append({
            "periode": parse_period(f["period"], "yyyyww"),
            "land": f["country"], "banner": f["banner"],
            "winkel_id": None, "winkel_naam": None,
            "merk": f["brand"],
            # De SKU is de identiteit, niet de GTIN. Geen enkel echt
            # DWH-bestand levert de GTIN-kolom, dus alle opgeslagen feiten
            # zijn al op SKU gekeyd (het per-winkel-gereedschap ook). Een
            # "GTIN als hij er is"-voorkeur zou bij één toekomstige export
            # mét die kolom elk artikel een nieuwe identiteit geven: alles
            # tegelijk NIEUW én DELISTED, en dubbeltellingen omdat de
            # hersleveringsvervanging de oude rijen niet meer herkent.
            "artikel_ean": f["unit"],
            "artikel_naam": info.get("article_description"),
            "volume": int(round(f["sales_volume"] or 0)),
            "omzet": float(f["sales_value"] or 0),
        })
    if not facts:
        raise ParseError("bestand bevat geen datarijen")
    # Eén rij per natuurlijke sleutel, fail-closed — zoals de Etos- en
    # ICI-parsers dat al doen. Dubbele weekkolommen zouden anders stil
    # dubbel tellen: de importer vervangt alleen wat er al STOND, niet
    # dubbelingen binnen dezelfde batch.
    keys = [(x["merk"], x["land"], x["banner"], x["artikel_ean"], x["periode"])
            for x in facts]
    if len(keys) != len(set(keys)):
        dubbel = sorted({k for k in keys if keys.count(k) > 1})[:5]
        raise ParseError(
            f"dubbele merk/artikel/week-combinaties in het bestand (o.a. "
            f"{dubbel}); import afgebroken om dubbeltelling te voorkomen")
    return {"facts": facts, "periode_type": "week",
            "periodes": sorted({x["periode"] for x in facts}),
            "warnings": parsed.warnings}


def parse_file(filename: str, content: bytes, profile: Profile) -> dict:
    """Parse + validate a whole file against a profile.

    Returns {'facts': [...], 'periode_type': .., 'periodes': [..]} or raises
    ParseError carrying row-level errors (PROMPT §3: never half imports).
    """
    d = profile.definition
    builtin = d.get("builtin")
    if builtin == "kruidvat_dwh":
        return _parse_builtin_kruidvat(filename, content)
    if builtin == "ici_maandrapport":
        from . import ici_maandrapport
        try:
            return ici_maandrapport.parse_workbook(
                content, land=d.get("constants", {}).get("land") or "NL")
        except ValueError as e:
            raise ParseError(str(e))
        except Exception as e:  # noqa: BLE001 - onleesbaar bestand is geen crash
            raise ParseError(f"bestand kon niet worden gelezen: {e}") from e
    if builtin == "etos_datagrid":
        from . import etos_datagrid
        try:
            return etos_datagrid.parse_workbook(content)
        except ValueError as e:
            raise ParseError(str(e))
        except Exception as e:  # noqa: BLE001 - onleesbaar bestand is geen crash
            raise ParseError(f"bestand kon niet worden gelezen: {e}") from e
    if builtin:
        raise ParseError(f"onbekende ingebouwde parser {builtin!r}")
    det = d["detection"]
    try:
        headers, data = read_table(content, det)
    except ParseError:
        raise
    except Exception as e:  # noqa: BLE001 - onleesbaar bestand is geen crash
        raise ParseError(f"bestand kon niet worden gelezen: {e}") from e

    # Twee kolommen met dezelfde kop maken de mapping dubbelzinnig: welke
    # van de twee bedoelt het profiel? Weigeren in plaats van gokken.
    folded = [h.casefold() for h in headers if h]
    dupes = sorted({h for h in folded if folded.count(h) > 1})
    if dupes:
        raise ParseError("dubbele kolomkoppen: " + ", ".join(dupes))
    hindex = {h.lower(): i for i, h in enumerate(headers)}

    missing_headers = [req for req in det.get("required_headers", [])
                       if req.lower() not in hindex]
    if missing_headers:
        raise ParseError("verplichte kolommen ontbreken: " + ", ".join(missing_headers))

    period_cfg = d["period"]
    period_col = period_cfg["source_column"].lower()
    if period_col not in hindex:
        raise ParseError(f"periodekolom {period_cfg['source_column']!r} niet gevonden")

    mapping = [(m["source"].lower(), m["target"], m.get("normalize"))
               for m in d.get("mapping", []) if m.get("target")]
    for source, target, _n in mapping:
        if source not in hindex:
            raise ParseError(f"gemapte kolom {source!r} niet gevonden in het bestand")

    caps = capabilities(d)
    decimal = det.get("decimal") or ","
    constants = d.get("constants", {})

    facts: list[dict] = []
    errors: list[dict] = []
    for rownum, row in enumerate(data, start=int(det.get("header_row") or 1) + 1):
        fact = {"land": None, "banner": None, "winkel_id": None, "winkel_naam": None,
                "merk": None, "artikel_ean": None, "artikel_naam": None}
        fact.update(constants)
        try:
            fact["periode"] = parse_period(row[hindex[period_col]], period_cfg["format"])
        except (PeriodError, IndexError) as e:
            errors.append({"rij": rownum, "veld": "periode", "fout": str(e)})
            continue
        row_ok = True
        for source, target, normalize in mapping:
            raw = row[hindex[source]] if hindex[source] < len(row) else ""
            try:
                if target in ("volume", "omzet"):
                    value = parse_number(raw, decimal)
                    if target == "volume":
                        # Stil afronden verandert de data zonder dat iemand
                        # het ziet, en round() rondt halven naar EVEN af: 10,5
                        # werd 10 maar 11,5 werd 12. Een aantal stuks hoort
                        # geheel te zijn; is het dat niet, dan klopt er iets
                        # aan de bron en moet de gebruiker dat weten.
                        if value != int(value):
                            raise ValueError(
                                f"aantal {value!r} is niet heel; volume telt stuks")
                        value = int(value)
                else:
                    value = str(raw).strip()
                    if normalize == "upper":
                        value = value.upper()
                    if target == "artikel_ean" and value and not EAN_RE.fullmatch(value):
                        raise ValueError(f"EAN {value!r} is niet 8 of 13 cijfers")
            except (ValueError, TypeError) as e:
                errors.append({"rij": rownum, "veld": target, "fout": str(e)})
                row_ok = False
                break
            fact[target] = value
        if not row_ok:
            continue
        for req in ("volume", "omzet"):
            if fact.get(req) is None:
                errors.append({"rij": rownum, "veld": req, "fout": "verplicht veld leeg"})
                row_ok = False
        if caps["merk"] and not fact.get("merk"):
            errors.append({"rij": rownum, "veld": "merk", "fout": "verplicht veld leeg"})
            row_ok = False
        if row_ok:
            facts.append(fact)

    if errors:
        raise ParseError(f"{len(errors)} rij(en) met fouten", row_errors=errors)
    if not facts:
        raise ParseError("bestand bevat geen datarijen")

    return {
        "facts": facts,
        "periode_type": period_cfg["type"],
        "periodes": sorted({f["periode"] for f in facts}),
    }
