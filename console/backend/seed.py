"""`make seed`: build the seed files in their REAL delivery format from the
CSV stand-ins in console/seed/, then push them through the actual import
pipeline (no direct inserts — this proves the parsers work, PROMPT §7).

Also loads the four handoff profiles verbatim, default settings and the
mock SharePoint contract documents.
"""

from __future__ import annotations

import csv
import io
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import db
from engine import contracts, importer

BASE = Path(__file__).resolve().parents[1]          # console/
PROFILES = BASE / "profiles"
SEED = BASE / "seed"


def _read_standin(path: Path) -> tuple[list[str], list[list[str]]]:
    rows = [r for r in csv.reader(path.open(), delimiter=";")
            if r and not r[0].startswith("#")]
    return rows[0], rows[1:]


def _to_xlsx(headers, data, sheet: str, header_row: int, numeric_cols: dict) -> bytes:
    """numeric_cols: {header: 'int'|'float'} — floats came in as '7.180,00'."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    for i in range(1, header_row):
        ws.cell(row=i, column=1,
                value=f"metadataregel {i}" if header_row > 2 else "")
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    for r, row in enumerate(data, start=header_row + 1):
        for c, (h, val) in enumerate(zip(headers, row), start=1):
            kind = numeric_cols.get(h)
            if kind == "int":
                val = int(val)
            elif kind == "float":
                val = float(val.replace(".", "").replace(",", "."))
            ws.cell(row=r, column=c, value=val)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_dwh_xlsx(rows: list[dict], weeks: list[str] | None = None,
                  country3="NLD", formula="KV", brand_meta=None) -> bytes:
    """Generate a workbook in the REAL Kruidvat DWH-export layout (metadata
    block, two-row header with side-by-side week Volume/Value pairs, Total
    row and trailing Total column), so seed and tests exercise the builtin
    parser exactly like a production file.

    rows: [{sku, gtin, desc, brand, weeks: {"202632": (vol, val), ...}}]
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if weeks is None:
        weeks = sorted({wk for r in rows for wk in r["weeks"]})
    brand_meta = brand_meta or ";".join(sorted({r["brand"] for r in rows}))
    for i, (label, value) in enumerate(
            [("Country:", country3), ("Formula:", formula), ("Brand:", brand_meta),
             ("Weeks:", len(weeks)), ("Date:", "2026-08-14")], start=1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
    h1, h2 = 7, 8
    for c, name in enumerate(["Brand", "GTIN/PLU", "SKU No.", "Article Description"], start=1):
        ws.cell(row=h1, column=c, value=name)
    col = 5
    week_cols = {}
    for wk in weeks:
        ws.cell(row=h1, column=col, value=wk)
        ws.cell(row=h2, column=col, value="Sales Volume")
        ws.cell(row=h2, column=col + 1, value="Sales Value")
        week_cols[wk] = col
        col += 2
    ws.cell(row=h1, column=col, value="Total")
    ws.cell(row=h2, column=col, value="Sales Volume")
    ws.cell(row=h2, column=col + 1, value="Sales Value")

    week_totals = {wk: 0 for wk in weeks}
    r = h2 + 1
    for row in rows:
        ws.cell(row=r, column=1, value=row["brand"])
        ws.cell(row=r, column=2, value=row["gtin"])
        ws.cell(row=r, column=3, value=int(row["sku"]))
        ws.cell(row=r, column=4, value=row["desc"])
        tot_vol = tot_val = 0
        for wk, (vol, val) in row["weeks"].items():
            ws.cell(row=r, column=week_cols[wk], value=vol)
            ws.cell(row=r, column=week_cols[wk] + 1, value=val)
            week_totals[wk] += vol
            tot_vol += vol
            tot_val += val
        ws.cell(row=r, column=col, value=tot_vol)
        ws.cell(row=r, column=col + 1, value=round(tot_val, 2))
        r += 1
    ws.cell(row=r, column=1, value="Total")
    ws.cell(row=r, column=3, value=len(rows))
    for wk, c in week_cols.items():
        ws.cell(row=r, column=c, value=week_totals[wk])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _kv_demo_rows(weeks: list[str], factor=1.0) -> list[dict]:
    items = [("31210001", "4049469072773", "Tweezerman Slant Tweezer", "TWEEZERMAN", 18.0),
             ("31210002", "4049469083120", "Tweezerman Nail Clipper", "TWEEZERMAN", 13.0),
             ("31210003", "4064089040111", "Striplac Rose", "ALESSANDRO", 14.0)]
    out = []
    for i, (sku, gtin, desc, brand, prijs) in enumerate(items):
        wkdata = {}
        for j, wk in enumerate(weeks):
            vol = 8 + (i * 3 + j) % 7
            wkdata[wk] = (vol, round(vol * prijs * factor, 2))
        out.append({"sku": sku, "gtin": gtin, "desc": desc, "brand": brand, "weeks": wkdata})
    return out


def make_ici_xlsx(blocks: dict[str, dict[str, dict[str, float]]]) -> bytes:
    """Genereer een werkboek in het ECHTE ICI-maandrapportformaat: een
    'Stores'-tab met merkblokken (Store/Address-kop + YYYYMM-kolommen) en
    een 'Brands'-tab met maandtotalen per merk voor de reconciliatie.

    blocks: {merk: {winkel_id: {"202501": omzet, ...}}}
    """
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Stores"
    months = sorted({m for stores in blocks.values() for w in stores.values() for m in w})
    r = 1
    for merk, stores in blocks.items():
        r += 1
        ws.cell(row=r, column=2, value=merk)
        r += 2
        ws.cell(row=r, column=3, value="Store")
        ws.cell(row=r, column=4, value="Address")
        for j, m in enumerate(months):
            ws.cell(row=r, column=5 + j, value=m)
        ws.cell(row=r, column=5 + len(months), value="Total")
        for store, per_month in stores.items():
            r += 1
            ws.cell(row=r, column=3, value=int(store))
            ws.cell(row=r, column=4, value=f"DEMOSTAD - winkel {store}")
            for j, m in enumerate(months):
                if m in per_month:
                    ws.cell(row=r, column=5 + j, value=per_month[m])
            ws.cell(row=r, column=5 + len(months), value=round(sum(per_month.values()), 2))
        r += 2

    wsb = wb.create_sheet("Brands")
    wsb.cell(row=2, column=2, value="Year")
    wsb.cell(row=2, column=3, value="Category")
    for m in range(1, 13):
        wsb.cell(row=2, column=3 + m, value=f"{m:02d}")
    rr = 3
    years = sorted({m[:4] for m in months})
    for year in years:
        for merk, stores in blocks.items():
            wsb.cell(row=rr, column=2, value=int(year))
            wsb.cell(row=rr, column=3, value=merk)
            for m in range(1, 13):
                key = f"{year}{m:02d}"
                total = round(sum(pm.get(key, 0) for pm in stores.values()), 2)
                if total:
                    wsb.cell(row=rr, column=3 + m, value=total)
            rr += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_etos_xlsx(artikelen: list[dict], weeks: list[str] | None = None,
                   brand_count: int | None = None) -> bytes:
    """Genereer een werkboek in het ECHTE Etos Data Grid-widgetformaat:
    metadatablok (Fiscal YTD-range + Brand (N)), weekkoppen met de
    ISO-zondag als Ending-datum, en de UPC/Brand-subkop met per week een
    Sales/Units-paar — zodat seed en tests de ingebouwde parser exact zo
    raken als een productie-export.

    artikelen: [{upc, naam, merk, merk_nr, weeks: {"202632": (sales, units)}}]
    """
    import datetime as _dt

    from openpyxl import Workbook
    if weeks is None:
        weeks = sorted({w for a in artikelen for w in a["weeks"]})
    merken = sorted({f'{a["merk"]} - {a["merk_nr"]}' for a in artikelen})
    wb = Workbook()
    ws = wb.active
    ws.title = "Data_Grid_57018_widget"
    ws.cell(row=4, column=1, value="File")
    ws.cell(row=5, column=1, value="File Generated")
    ws.cell(row=5, column=2, value="2026-08-17 19:54 CEST")
    ws.cell(row=6, column=1, value="Board Name")
    ws.cell(row=6, column=2, value="Sales per item analyse")
    ws.cell(row=7, column=1, value="Widget Name")
    ws.cell(row=7, column=2, value="Data Grid")
    n_merken = brand_count if brand_count is not None else len(merken)
    ws.cell(row=11, column=1, value="Product :")
    ws.cell(row=11, column=2, value=f"Brand ({n_merken})")
    ws.cell(row=18, column=1, value='Time "Fiscal YTD"')
    ws.cell(row=18, column=2,
            value=f"Fiscal YTD {weeks[0]}-{weeks[-1]}, Ending {weeks[-1]}")
    ws.cell(row=19, column=1, value=f"Brand ({n_merken})")
    ws.cell(row=19, column=2, value=", ".join(merken))
    for j, wk in enumerate(weeks):
        jaar, week = int(wk[:4]), int(wk[4:])
        zondag = _dt.date.fromisocalendar(jaar, week, 7)
        ws.cell(row=20, column=4 + j * 2,
                value=f"{wk} (Ending {zondag.strftime('%d/%m/%Y')})")
        ws.cell(row=21, column=4 + j * 2, value="Sales € TY")
        ws.cell(row=21, column=5 + j * 2, value="Units TY")
    ws.cell(row=21, column=1, value="UPC Name")
    ws.cell(row=21, column=2, value="UPC ID")
    ws.cell(row=21, column=3, value="Brand")
    r = 21
    for a in artikelen:
        r += 1
        ws.cell(row=r, column=1, value=a["naam"])
        ws.cell(row=r, column=2, value=str(a["upc"]))
        ws.cell(row=r, column=3, value=f'{a["merk"]} - {a["merk_nr"]}')
        for j, wk in enumerate(weeks):
            if wk in a["weeks"]:
                sales, units = a["weeks"][wk]
                ws.cell(row=r, column=4 + j * 2, value=sales)
                ws.cell(row=r, column=5 + j * 2, value=units)
    ws.cell(row=r + 5, column=1,
            value="All insights and reports are for internal use only.")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _etos_demo_artikelen(weeks: list[str], factor: float = 1.0) -> list[dict]:
    basis = [("120789202", "BJORN AXEN VOLUMIZING SHAMPOO", "BJÖRN AXÉN", "2912", 11.5),
             ("120727773", "TWEEZERMAN LASH CURLER", "TWEEZERMAN", "2278", 19.75),
             ("120742623", "PATCHOLOGY FLASHPATCH EYE GELS", "PATCHOLOGY", "2731", 13.25)]
    out = []
    for i, (upc, naam, merk, nr, prijs) in enumerate(basis):
        wkdata = {}
        for k, wk in enumerate(weeks):
            units = int((8 + (i * 7 + k * 3) % 21) * factor)
            wkdata[wk] = (round(units * prijs, 2), units)
        out.append({"upc": upc, "naam": naam, "merk": merk, "merk_nr": nr,
                    "weeks": wkdata})
    return out


def _ici_demo_blocks(months: list[str]) -> dict:
    blocks: dict = {}
    for merk, base in (("TWEEZERMAN", 60.0), ("DEPEND", 22.0)):
        stores: dict = {}
        for s, store in enumerate(("6051", "6053", "6054")):
            stores[store] = {m: round(base + (s * 7 + i * 3) % 25, 2)
                             for i, m in enumerate(months)}
        blocks[merk] = stores
    return blocks


def build_seed_files() -> list[tuple[str, bytes]]:
    files = []

    files.append(("DWH__Sales_volume__sales_Tweezerman_KVNL_32_demo.xlsx",
                  make_dwh_xlsx(_kv_demo_rows(["202632"]))))

    files.append(("Maandelijkse_resultaten__Tweezerman__Depend_ICI_Paris_XL__demo.xlsx",
                  make_ici_xlsx(_ici_demo_blocks(["202607"]))))

    files.append(("Data_Grid_57018_widget_demo.xlsx",
                  make_etos_xlsx(_etos_demo_artikelen(
                      [f"2026{wk:02d}" for wk in range(30, 33)]))))

    h, d = _read_standin(SEED / "douglas_Abverkauf_KW32.csv")
    files.append(("Douglas_Abverkauf_KW32.xlsx",
                  _to_xlsx(h, d, "Sheet1", 1,
                           {"Absatz": "int", "Umsatz": "float"})))
    return files


# VERZONNEN demo-instellingen. Bewust NIET onderdeel van bootstrap(): een
# verzonnen winkelaantal of rotatietarget voedt echte berekeningen (omzet per
# winkel, delist-advies) en levert dan geloofwaardige maar onjuiste cijfers.
# Een verse installatie start met lege instellingen die de gebruiker zelf
# invult; deze waarden komen alleen mee met `make seed` (demo).
DEMO_SETTINGS = {
    "kruidvat": {
        "winkels_targets": [
            ("TWEEZERMAN", "NL", "KV", 912, 45.0),
            ("TWEEZERMAN", "NL", "TP", 410, 30.0),
            ("TWEEZERMAN", "BE", "KV", 214, 40.0),
        ],
        "rotatie": [("TWEEZERMAN", 8.0), ("ALESSANDRO", 6.0)],
        "mail": [("DWH weeklevering", "rapportage@kruidvat.nl", "DWH_sellout_*.xlsx")],
    },
    "ici-paris-xl": {
        # aantal_winkels NULL: comes from the facts (readonly in the UI)
        "winkels_targets": [("TWEEZERMAN", "NL", None, None, 1500.0),
                            ("TWEEZERMAN", "BE", None, None, 1500.0),
                            ("ALESSANDRO", "NL", None, None, 900.0),
                            ("ALESSANDRO", "BE", None, None, 900.0)],
        "rotatie": [],
        "mail": [("ICI maandlevering", "reports@iciparisxl.be", "*ICI_Paris*.xlsx")],
    },
}


def build_history_files() -> list[tuple[str, bytes]]:
    """Deterministic demo history (2024-2026) in each retailer's real
    format, so trend/YTD/sparklines have something to show. Goes through
    the same import pipeline as everything else."""
    import math

    files = []
    kv_items = [("4049469072773", "Tweezerman Slant Tweezer", 18.0),
                ("4049469083120", "Tweezerman Nail Clipper", 13.0),
                ("4064089040111", "Striplac Rose", 14.0)]

    def wave(base, year, week, amp=0.25):
        seasonal = 1 + amp * math.sin((week / 52) * 2 * math.pi + year)
        growth = 1 + (year - 2024) * 0.08
        return base * seasonal * growth

    for year in (2024, 2025, 2026):
        weeks = range(1, 33 if year == 2026 else 53)
        week_labels = [f"{year}{wk:02d}" for wk in weeks]
        factor = 1 + (year - 2024) * 0.08
        files.append((f"DWH__Sales_volume__sales_Tweezerman_KVNL_{year}_demo.xlsx",
                      make_dwh_xlsx(_kv_demo_rows(week_labels, factor=factor))))


        month_labels = [f"{year}{m:02d}" for m in range(1, 8 if year == 2026 else 13)]
        files.append((f"Maandelijkse_resultaten__Tweezerman__Depend_ICI_Paris_XL__{year}.xlsx",
                      make_ici_xlsx(_ici_demo_blocks(month_labels))))

        files.append((f"Data_Grid_{57000 + year - 2024}_widget.xlsx",
                      make_etos_xlsx(_etos_demo_artikelen(week_labels, factor=factor))))
    return files


def bootstrap():
    """Wat een verse installatie nodig heeft om te WERKEN: alleen de
    parser-profielen. Geen instellingen, geen contracten, geen cijfers —
    alles wat verzonnen is, hoort in seed() (demo). Idempotent; draait
    bij ELKE containerstart, zodat een nieuw meegeleverd profiel ook op
    een bestaande installatie aankomt."""
    db.init_db()
    with db.get_conn() as conn:
        for path in sorted(PROFILES.glob("*.json")):
            d = json.loads(path.read_text())
            conn.execute(
                "INSERT OR IGNORE INTO parser_profiles (retailer_id, version, status, definition, "
                "published_at) VALUES (?,?,?,?, CASE WHEN ?='live' THEN datetime('now') END)",
                (d["retailer_id"], d["version"], d["status"],
                 json.dumps(d, ensure_ascii=False), d["status"]))
            if d["status"] == "live":
                # Zonder deze vlag blijft de retailerkaart op "NIEUW" staan
                # terwijl er wel degelijk een werkende parser is.
                conn.execute("UPDATE retailers SET aangesloten=1 WHERE id=?",
                             (d["retailer_id"],))


def _load_demo_settings():
    """Verzonnen instellingen + contractdocumenten — alleen voor de demo."""
    with db.get_conn() as conn:
        for retailer, cfg in DEMO_SETTINGS.items():
            conn.executemany(
                "INSERT OR IGNORE INTO retailer_settings (retailer_id, merk, land, banner, "
                "aantal_winkels, target_per_winkel) VALUES (?,?,?,?,?,?)",
                [(retailer, *row) for row in cfg["winkels_targets"]])
            conn.executemany(
                "INSERT OR REPLACE INTO rotatie_targets (retailer_id, merk, stuks_per_winkel_per_week) "
                "VALUES (?,?,?)", [(retailer, *row) for row in cfg["rotatie"]])
            for naam, afzender, glob in cfg["mail"]:
                if not conn.execute("SELECT 1 FROM mail_rules WHERE retailer_id=? AND naam=?",
                                    (retailer, naam)).fetchone():
                    conn.execute(
                        "INSERT INTO mail_rules (retailer_id, naam, afzender, bijlage_glob) "
                        "VALUES (?,?,?,?)", (retailer, naam, afzender, glob))

        # SharePoint: link + interpret documents for retailers present in contracts.json.
        doclist = json.loads((SEED / "contracts.json").read_text())
        for retailer in doclist:
            conn.execute(
                "INSERT INTO sharepoint_links (retailer_id, map_url) VALUES (?,?) "
                "ON CONFLICT(retailer_id) DO UPDATE SET map_url=excluded.map_url",
                (retailer, f"https://leijtens.sharepoint.com/sites/retail/{retailer}/contracten"))
            contracts.sync_documents(conn, retailer)


def seed():
    """bootstrap() + verzonnen instellingen + demodata door de ECHTE
    import-pipeline. Uitsluitend voor demonstratie."""
    bootstrap()
    _load_demo_settings()
    with db.get_conn() as conn:
        results = [importer.run_import(conn, name, content)
                   for name, content in build_seed_files()]
        history = [importer.run_import(conn, name, content)
                   for name, content in build_history_files()]

    for r in results + history:
        print(f"[seed] {r['filename']}: {r['status']} "
              f"({r['rows']} rijen, retailer={r['retailer_id']})")
    assert all(h["status"] in ("ingelezen", "test") for h in history), history
    statuses = {r["filename"]: r["status"] for r in results}
    assert statuses["DWH__Sales_volume__sales_Tweezerman_KVNL_32_demo.xlsx"] == "ingelezen", statuses
    assert statuses["Maandelijkse_resultaten__Tweezerman__Depend_ICI_Paris_XL__demo.xlsx"] == "ingelezen", statuses
    assert statuses["Douglas_Abverkauf_KW32.xlsx"] == "profiel_nodig", statuses
    print("[seed] klaar — alle verwachte statussen kloppen")


if __name__ == "__main__":
    seed()
