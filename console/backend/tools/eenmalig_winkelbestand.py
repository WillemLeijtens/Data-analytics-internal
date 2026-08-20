"""Eenmalig een Kruidvat-export OP WINKELNIVEAU inlezen.

Waarom een los script en niet de parser? De reguliere DWH-export heeft de
metadata op rij 1-5 en de kop op rij 7/8. Deze variant
(`..._per_week_per_store_...`) heeft de metadata verspreid over rij 1-20 met
de waarde in kolom C, de kop op rij 22/23, en extra kolommen voor winkel,
plaats en voorraad. De ingebouwde parser wijst hem daarom af. Dit script leest
die ene indeling, zonder aan de parser te komen.

Het staat bewust in tools/ en niet in engine/: het hoort niet bij de
importketen en wordt met de hand gedraaid.

Twee dingen die het veilig maken:

  * Weken die al GEAGGREGEERD in de database staan (winkel_id leeg) worden
    overgeslagen. De feitensleutel bevat winkel_id, dus winkelrijen zouden
    er anders naast komen te staan in plaats van eroverheen — en dan telt het
    dashboard beide mee.
  * Zonder --schrijf gebeurt er niets: dan toont het script alleen wat het
    zou doen.

Gebruik:
    python tools/eenmalig_winkelbestand.py <bestand.xlsx>             # kijken
    python tools/eenmalig_winkelbestand.py <bestand.xlsx> --schrijf   # doen
    python tools/eenmalig_winkelbestand.py <bestand.xlsx> --verwijder # ongedaan maken
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import db                                                  # noqa: E402
from engine.importer import _replace_redelivered_facts, file_hash  # noqa: E402
from engine.periods import PeriodError, parse_period        # noqa: E402

RETAILER_ID = "kruidvat"
BLAD = "Sales per week"
KOP_SKU = "SKU no."
KOP_WINKEL = "Store"
VOLUME, WAARDE = "Sales volume", "Gross sales value"
LAND = {"BEL": "BE", "NLD": "NL"}


class Afgebroken(RuntimeError):
    """Iets klopt niet aan het bestand; er wordt niets geschreven."""


def _tekst(waarde) -> str | None:
    """Celwaarde als nette tekst. Winkelnummers en SKU's komen als float uit
    Excel (6369354.0); die punt-nul hoort niet in een sleutel thuis."""
    if waarde is None:
        return None
    if isinstance(waarde, float) and waarde.is_integer():
        return str(int(waarde))
    tekst = str(waarde).strip()
    return tekst or None


def _metadata(ws) -> dict:
    """Land en formule uit het blok bovenaan.

    De labels staan in kolom B, de waarde in kolom C — maar door
    samengevoegde cellen staat die waarde soms één rij hoger dan het label.
    Beide plekken worden daarom bekeken."""
    gevonden: dict[str, str] = {}
    rijen = list(ws.iter_rows(min_row=1, max_row=20, max_col=4, values_only=True))
    for i, rij in enumerate(rijen):
        label = _tekst(rij[1]) if len(rij) > 1 else None
        if not label or not label.endswith(":"):
            continue
        naam = label.rstrip(":").strip()
        for bron in (rij, rijen[i - 1] if i else None):
            if bron is None:
                continue
            waarde = _tekst(bron[2]) if len(bron) > 2 else None
            if waarde:
                gevonden.setdefault(naam, waarde)
                break
    return gevonden


def _kolommen(ws) -> tuple[int, dict, dict]:
    """(rij waar de data begint, vaste kolommen, weekkolommen).

    De kop is tweeregelig: bovenin staat het weeknummer, eronder of het om
    volume of waarde gaat. Een weeknummer geldt tot het volgende."""
    rijen = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
    kop_i = None
    for i, rij in enumerate(rijen):
        waarden = {_tekst(c) for c in rij}
        if KOP_SKU in waarden and KOP_WINKEL in waarden:
            kop_i = i
            break
    if kop_i is None or kop_i + 1 >= len(rijen):
        raise Afgebroken(
            f"geen kopregel met '{KOP_SKU}' en '{KOP_WINKEL}' gevonden — "
            "is dit wel de export op winkelniveau?")

    boven, onder = rijen[kop_i], rijen[kop_i + 1]
    vast: dict[str, int] = {}
    weken: dict[str, dict[str, int]] = {}
    huidig = None
    for j in range(len(boven)):
        kop = _tekst(boven[j])
        if kop:
            huidig = kop
            if not kop.isdigit():
                vast.setdefault(kop, j)
        sub = _tekst(onder[j]) if j < len(onder) else None
        if huidig and huidig.isdigit() and sub in (VOLUME, WAARDE):
            weken.setdefault(huidig, {})[sub] = j

    ontbreekt = [k for k in (KOP_SKU, "Brand", KOP_WINKEL) if k not in vast]
    if ontbreekt:
        raise Afgebroken(f"kolommen ontbreken in de kop: {', '.join(ontbreekt)}")
    if not weken:
        raise Afgebroken("geen weekkolommen gevonden")
    return kop_i + 3, vast, weken       # +3: kop, subkop, en 1-gebaseerd tellen


def lees_bestand(pad: str) -> tuple[list[dict], dict]:
    """(feiten, samenvatting). Leest alleen het weekblad; 'Sales per day'
    bevat dezelfde omzet per dag en zou dus dubbeltellen."""
    bron = Path(pad)
    if not bron.is_file():
        raise Afgebroken(
            f"bestand niet gevonden: {pad}\n"
            "  Staat het al op de droplet? Twee manieren om het daar te "
            "krijgen:\n"
            "   1. upload het via het IMPORTSCHERM van de console. De parser "
            "wijst het af,\n"
            "      maar het bestand wordt bewaard in /srv/data/inbox/ — het "
            "pad staat in de\n"
            "      melding op het scherm. Alleen dit werkt vanuit een "
            "webconsole.\n"
            "   2. vanaf een eigen terminal: "
            "scp <bestand>.xlsx root@<droplet>:~/analytics/console/data/\n"
            "  Kijken wat er staat: ls -la ~/analytics/console/data/ "
            "~/analytics/console/data/inbox/")
    try:
        wb = openpyxl.load_workbook(pad, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001 - alles wat openpyxl niet lust
        raise Afgebroken(f"{bron.name} is geen leesbaar xlsx-bestand ({e})") from e
    if BLAD not in wb.sheetnames:
        raise Afgebroken(f"blad '{BLAD}' ontbreekt; gevonden: {wb.sheetnames}")
    ws = wb[BLAD]

    meta = _metadata(ws)
    land3 = meta.get("Country")
    banner = meta.get("Formula")
    if not land3 or not banner:
        raise Afgebroken(f"land of formule ontbreekt in de metadata: {meta}")
    land = LAND.get(land3.upper(), land3.upper()[:2])

    start, vast, weken = _kolommen(ws)
    k_sku, k_merk, k_winkel = vast[KOP_SKU], vast["Brand"], vast[KOP_WINKEL]
    k_naam = vast.get("Articledescription")
    k_plaats = vast.get("City")

    try:
        periode_van = {w: parse_period(w, "yyyyww") for w in weken}
    except PeriodError as e:
        raise Afgebroken(f"onbruikbaar weeknummer in de kop: {e}") from e

    feiten: list[dict] = []
    winkels: set[str] = set()
    per_merk: dict[str, list] = defaultdict(lambda: [0.0, 0.0])
    for rij in ws.iter_rows(min_row=start, values_only=True):
        merk = _tekst(rij[k_merk]) if k_merk < len(rij) else None
        sku = _tekst(rij[k_sku]) if k_sku < len(rij) else None
        winkel = _tekst(rij[k_winkel]) if k_winkel < len(rij) else None
        if not merk or not sku or not winkel:
            continue                    # lege regel of een totaalregel
        winkels.add(winkel)
        for week, kolommen in weken.items():
            volume = rij[kolommen[VOLUME]] if VOLUME in kolommen and kolommen[VOLUME] < len(rij) else None
            omzet = rij[kolommen[WAARDE]] if WAARDE in kolommen and kolommen[WAARDE] < len(rij) else None
            if not volume and not omzet:
                continue
            feiten.append({
                "periode_type": "week", "periode": periode_van[week],
                "land": land, "banner": banner,
                "winkel_id": winkel,
                "winkel_naam": _tekst(rij[k_plaats]) if k_plaats is not None and k_plaats < len(rij) else None,
                "merk": merk, "artikel_ean": sku,
                "artikel_naam": _tekst(rij[k_naam]) if k_naam is not None and k_naam < len(rij) else None,
                "volume": float(volume) if volume else 0.0,
                "omzet": float(omzet) if omzet else 0.0,
            })
            per_merk[merk][0] += float(volume or 0)
            per_merk[merk][1] += float(omzet or 0)

    if not feiten:
        raise Afgebroken("geen enkele regel met omzet of volume gevonden")
    samenvatting = {
        "land": land, "banner": banner,
        "winkels": len(winkels),
        "weken": sorted({f["periode"] for f in feiten}),
        "per_merk": {m: {"volume": v[0], "omzet": v[1]} for m, v in sorted(per_merk.items())},
    }
    return feiten, samenvatting


def bestaande_periodes(conn, feiten: list[dict]) -> set:
    """(merk, land, banner, periode) die al GEAGGREGEERD in de database staan.

    Alleen rijen zonder winkel_id tellen: die zijn per week opgeteld en zouden
    naast de winkelrijen blijven staan."""
    scopes = {(f["merk"], f["land"], f["banner"], f["periode"]) for f in feiten}
    bestaand = set()
    for merk, land, banner, periode in scopes:
        rij = conn.execute(
            "SELECT 1 FROM sellout_facts WHERE retailer_id=? AND merk=? AND land=? "
            "AND COALESCE(banner,'')=? AND periode=? AND winkel_id IS NULL LIMIT 1",
            (RETAILER_ID, merk, land, banner or "", periode)).fetchone()
        if rij:
            bestaand.add((merk, land, banner, periode))
    return bestaand


def schrijf(conn, pad: str, feiten: list[dict]) -> int:
    """Feiten wegschrijven in één transactie, herhaalbaar.

    Dezelfde sleutelvervanging als de gewone import, plus een regel in
    `imports` zodat het scherm Importstatus laat zien dat dit gebeurd is."""
    # Dezelfde regel als de gewone import: een bestand dat opnieuw wordt
    # aangeboden vervangt zijn vorige lading in plaats van ernaast te komen.
    # Zonder dit botst de tweede keer op de unieke file_hash in `imports`.
    hash_ = file_hash(Path(pad).read_bytes())
    vorige = conn.execute("SELECT id FROM imports WHERE file_hash=?", (hash_,)).fetchone()
    if vorige:
        conn.execute("DELETE FROM sellout_facts WHERE import_id=?", (vorige["id"],))
        conn.execute("DELETE FROM imports WHERE id=?", (vorige["id"],))

    _replace_redelivered_facts(conn, RETAILER_ID, feiten)
    periodes = sorted({f["periode"] for f in feiten})
    cur = conn.execute(
        "INSERT INTO imports (retailer_id, profile_id, filename, file_hash, periode_type, "
        "periode, row_count, status, error_detail) VALUES (?,?,?,?,?,?,?,?,?)",
        (RETAILER_ID, None, Path(pad).name, hash_, "week",
         f"{periodes[0]} t/m {periodes[-1]} ({len(periodes)})", len(feiten), "ingelezen",
         json.dumps({"warnings": [
             "eenmalig ingelezen met tools/eenmalig_winkelbestand.py "
             "(export op winkelniveau; de reguliere parser kent deze indeling niet)"]},
             ensure_ascii=False)))
    import_id = cur.lastrowid
    conn.executemany(
        "INSERT INTO sellout_facts (retailer_id, import_id, periode_type, periode, land, "
        "banner, winkel_id, winkel_naam, merk, artikel_ean, artikel_naam, volume, omzet) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [(RETAILER_ID, import_id, f["periode_type"], f["periode"], f["land"], f["banner"],
          f["winkel_id"], f["winkel_naam"], f["merk"], f["artikel_ean"], f["artikel_naam"],
          f["volume"], f["omzet"]) for f in feiten])
    return import_id


def verwijder(conn, pad: str) -> tuple[int, int]:
    """Alles wat dit bestand ooit heeft weggeschreven weer weg.

    Op de file_hash, dus precies de lading van dít bestand — feiten van
    andere imports blijven staan. Nodig omdat winkelrijen de analyses
    merkbaar trager maken: 54.000 extra regels maakten het dashboard bij
    Kruidvat ruim zeven keer zo traag."""
    hash_ = file_hash(Path(pad).read_bytes())
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM imports WHERE file_hash=?", (hash_,))]
    rijen = 0
    for i in ids:
        rijen += conn.execute(
            "DELETE FROM sellout_facts WHERE import_id=?", (i,)).rowcount
        conn.execute("DELETE FROM imports WHERE id=?", (i,))
    return len(ids), rijen


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("bestand")
    p.add_argument("--schrijf", action="store_true",
                   help="daadwerkelijk wegschrijven (zonder deze vlag alleen tonen)")
    p.add_argument("--verwijder", action="store_true",
                   help="de eerder ingelezen lading van dit bestand weer weghalen")
    a = p.parse_args(argv)

    if a.verwijder:
        if not Path(a.bestand).is_file():
            print(f"Afgebroken: bestand niet gevonden: {a.bestand}", file=sys.stderr)
            return 1
        with db.get_conn() as conn:
            imports, rijen = verwijder(conn, a.bestand)
        if not imports:
            print("Niets te verwijderen: dit bestand is hier nooit ingelezen.")
        else:
            print(f"Verwijderd: {rijen} feiten uit {imports} import(s). "
                  "De rest van de database is ongemoeid.")
        return 0

    try:
        feiten, s = lees_bestand(a.bestand)
    except Afgebroken as e:
        print(f"Afgebroken: {e}", file=sys.stderr)
        return 1

    eur = lambda v: f"€ {v:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
    print(f"Bestand   : {Path(a.bestand).name}")
    print(f"Retailer  : {RETAILER_ID}  land {s['land']}  formule {s['banner']}")
    print(f"Weken     : {s['weken'][0]} t/m {s['weken'][-1]} ({len(s['weken'])})")
    print(f"Winkels   : {s['winkels']}")
    for merk, w in s["per_merk"].items():
        print(f"  {merk:16s} volume {w['volume']:10,.0f}   omzet {eur(w['omzet'])}")

    with db.get_conn() as conn:
        klaar = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='sellout_facts'"
        ).fetchone()
        if not klaar:
            print("\nAfgebroken: deze database is nog niet ingericht. Draai dit "
                  "script in de console-container, waar CONSOLE_DB naar de echte "
                  "database wijst:\n  docker compose exec console python "
                  "tools/eenmalig_winkelbestand.py /srv/data/<bestand>.xlsx",
                  file=sys.stderr)
            return 1
        bestaand = bestaande_periodes(conn, feiten)
        if bestaand:
            weken = sorted({p for *_, p in bestaand})
            over = [f for f in feiten
                    if (f["merk"], f["land"], f["banner"], f["periode"]) in bestaand]
            print(f"\nOvergeslagen: {len(weken)} week(en) staan al geaggregeerd in de "
                  f"database — {', '.join(weken)}")
            print(f"              {len(over)} regels, {eur(sum(f['omzet'] for f in over))}")
            print("              (winkelrijen zouden daar naast de bestaande regels "
                  "komen en dus dubbel meetellen)")
            feiten = [f for f in feiten
                      if (f["merk"], f["land"], f["banner"], f["periode"]) not in bestaand]
        if not feiten:
            print("\nNiets over om in te lezen.")
            return 0

        weken_nieuw = sorted({f["periode"] for f in feiten})
        print(f"\nIn te lezen : {len(feiten)} regels, {len(weken_nieuw)} weken "
              f"({weken_nieuw[0]} t/m {weken_nieuw[-1]}), "
              f"{eur(sum(f['omzet'] for f in feiten))}")
        if not a.schrijf:
            print("\nProefronde — er is niets gewijzigd. Draai opnieuw met --schrijf.")
            return 0
        import_id = schrijf(conn, a.bestand, feiten)
    print(f"\nKlaar. Import {import_id} weggeschreven; zichtbaar bij Importstatus.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
