"""Kruidvat/Trekpleister parser: DWH sellout export files (.xlsx) into a
normalized long-format fact table plus item attributes. Weekly cadence,
SKU-grain (fact ``unit`` = SKU, ``period`` = ISO YYYYWW).

Real export format (confirmed against sample files):
  - Rows ~1-7: metadata block with labelled cells ("Country:", "Formula:",
    "Weeks:", "Brand:", "SKU no.:", "Date:") in some column, value in the
    next column.
  - Multiple sheets can exist; the authoritative one is the one whose data
    row count matches the SKU-count printed in its own trailing "Total"
    row (other sheets may repeat each SKU per GTIN/PLU variant).
  - A two-row column header: item attribute columns, then paired
    "Sales Volume" / "Sales Value" columns per year-week (format YYYYWW),
    with the week label only present on the first cell of each pair
    (merged cell) and needing forward-fill.
  - A trailing "Total" row whose SKU cell holds a row count, not a SKU.
"""

from __future__ import annotations

import math
import re

import openpyxl

from parsers.base import ParsedFile

RETAILER_ID = "KRUIDVAT"

LABELS = {"Country:", "Formula:", "Weeks:", "Brand:", "SKU no.:", "Date:"}
BANNER_NAMES = {"KV": "Kruidvat", "TP": "Trekpleister"}

FILENAME_RE = re.compile(
    r"([A-Za-z]+)_([A-Z]{2})([A-Z]{2})_(\d+)", re.IGNORECASE
)


def _cell_metadata(ws) -> dict:
    meta = {}
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value in LABELS:
                value_cell = ws.cell(row=cell.row, column=cell.column + 1)
                meta[cell.value.rstrip(":")] = value_cell.value
    return meta


def _find_header_rows(ws) -> tuple[int, int]:
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value == "SKU No.":
                return cell.row, cell.row + 1
    raise ValueError("Could not locate 'SKU No.' header row")


def _week_columns(ws, header_row1: int, header_row2: int, max_col: int):
    """Return list of (year_week, volume_col, value_col).

    The header row's week label is only set on the first cell of each
    merged Volume/Value pair, so a valid week label must be forward-filled
    across the pair — but a *non-week* label (e.g. a trailing "Total"
    column giving a per-SKU grand total across all weeks) must invalidate
    the carried-forward week rather than being mistaken for a repeat of
    the previous week.
    """
    weeks = []
    current_week = None
    col = 1
    while col <= max_col:
        top = ws.cell(row=header_row1, column=col).value
        if top is not None:
            if re.fullmatch(r"\d{6}", str(top)):
                current_week = str(top)
            else:
                current_week = None  # non-week header (e.g. "Total" column)
        sub = ws.cell(row=header_row2, column=col).value
        if sub == "Sales Volume" and current_week is not None:
            vol_col = col
            val_col = None
            nxt = ws.cell(row=header_row2, column=col + 1).value
            if nxt == "Sales Value":
                val_col = col + 1
            weeks.append((current_week, vol_col, val_col))
        col += 1
    return weeks


def _attribute_columns(ws, header_row1: int, header_row2: int, first_week_col: int):
    """Map known item-attribute header labels to their column index,
    restricted to columns before the first week column."""
    cols = {}
    for c in range(1, first_week_col):
        label = ws.cell(row=header_row1, column=c).value
        sub = ws.cell(row=header_row2, column=c).value
        name = label or sub
        if name:
            name = str(name).strip()
            cols.setdefault(name, c)  # first occurrence wins (e.g. two "Size" columns)
    return cols


def _sheet_candidates(wb):
    """Scan every worksheet that looks like a data sheet, returning
    (ws, header_row1, header_row2, n_data_rows, expected_total, total_row)."""
    candidates = []
    for ws in wb.worksheets:
        try:
            header_row1, header_row2 = _find_header_rows(ws)
        except ValueError:
            continue
        data_start = header_row2 + 1
        total_row = None
        n_data_rows = 0
        for r in range(data_start, ws.max_row + 1):
            desc_like = None
            sku_cell = ws.cell(row=r, column=3).value  # 'SKU No.' column is col C
            is_total = any(
                ws.cell(row=r, column=c).value == "Total"
                for c in range(1, 8)
            )
            if is_total:
                total_row = r
                break
            if sku_cell is not None:
                n_data_rows += 1
        expected = None
        if total_row is not None:
            expected = ws.cell(row=total_row, column=3).value
        candidates.append((ws, header_row1, header_row2, n_data_rows, expected, total_row))
    return candidates


def _pick_authoritative(candidates):
    """Among sheets describing the SAME feed, pick the one whose data-row
    count matches its own Total row — the others repeat each SKU per
    GTIN/PLU variant and would double-count."""
    for ws, h1, h2, n_data_rows, expected, total_row in candidates:
        # A malformed Total cell (text instead of the row count) must make
        # this sheet fail the match, not crash the whole parse.
        try:
            matches = expected is not None and int(expected) == n_data_rows
        except (TypeError, ValueError):
            matches = False
        if matches:
            return ws, h1, h2, total_row
    # fallback: first parseable sheet
    ws, h1, h2, _, _, total_row = candidates[0]
    return ws, h1, h2, total_row


def parse_filename(filename: str) -> dict | None:
    m = FILENAME_RE.search(filename)
    if not m:
        return None
    brand, banner, country, _num = m.groups()
    return {"brand": brand.upper(), "banner": banner.upper(), "country": country.upper()}


def _country_to_iso2(country3: str) -> str:
    mapping = {"BEL": "BE", "NLD": "NL"}
    return mapping.get(country3.upper(), country3.upper()[:2])


def _to_number(raw):
    """Coerce a cell value to float, or None if it isn't numeric. Without
    this, a text cell (e.g. 'n/a' or a stray '-') would flow straight into
    the sales_volume/sales_value REAL columns as a string — SQLite stores it
    happily, and pandas aggregations over the column then misbehave."""
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        number = float(raw)
        return number if math.isfinite(number) else None
    s = (str(raw).strip().replace("\u20ac", "").replace("\u00a0", "")
         .replace("\u202f", "").replace(" ", "").replace("'", ""))
    if "," in s and "." in s:
        decimal = "," if s.rfind(",") > s.rfind(".") else "."
        thousands = "." if decimal == "," else ","
        s = s.replace(thousands, "").replace(decimal, ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        number = float(s)
        return number if math.isfinite(number) else None
    except ValueError:
        return None


def parse_workbook(path: str, filename: str) -> ParsedFile:
    """Parse every distinct feed in the workbook.

    A workbook can hold several sheets. Two cases, told apart by each
    sheet's own metadata block (Brand/Country/Formula):

      * Same metadata on several sheets — one feed exported at different
        grains (the second repeats each SKU per GTIN/PLU variant). Only the
        authoritative sheet is used, or the figures would double-count.
      * DIFFERENT metadata — genuinely separate feeds in one file, e.g. a
        store banner and an "online" banner. Every one of them is parsed;
        an earlier version read only the first sheet's metadata and only one
        sheet, silently dropping the others (and their banner never reached
        the filters).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    candidates = _sheet_candidates(wb)
    if not candidates:
        raise ValueError("No parseable sheet found in workbook")

    # Group sheets by the feed their own metadata block describes.
    feeds: dict[tuple, list] = {}
    for cand in candidates:
        m = _cell_metadata(cand[0])
        key = (
            str(m.get("Brand", "")).strip(),
            str(m.get("Country", "")).strip(),
            str(m.get("Formula", "")).strip(),
        )
        feeds.setdefault(key, []).append(cand)

    primary_brand, primary_country3, primary_banner = next(iter(feeds))
    result = ParsedFile(
        retailer=RETAILER_ID,
        brand=primary_brand,
        country=_country_to_iso2(primary_country3),
        banner=primary_banner,
        source_filename=filename,
    )

    fn_info = parse_filename(filename)
    if fn_info:
        primary_country = _country_to_iso2(primary_country3)
        if fn_info["country"] != primary_country:
            result.warnings.append(
                f"Filename country '{fn_info['country']}' does not match "
                f"in-file country '{primary_country}' ({primary_country3})."
            )
        if fn_info["banner"] != primary_banner:
            result.warnings.append(
                f"Filename banner '{fn_info['banner']}' does not match "
                f"in-file formula '{primary_banner}'."
            )

    if len(feeds) > 1:
        listed = ", ".join(f"{b}/{_country_to_iso2(c)}/{bn}" for b, c, bn in feeds)
        result.warnings.append(
            f"Workbook contains {len(feeds)} feeds — all imported: {listed}."
        )

    for (brand, country3, banner), group in feeds.items():
        _parse_feed(result, group, brand, _country_to_iso2(country3), banner, country3)

    # The metadata block names the brand(s) the export was *requested* for and
    # joins several with ';' ("ALESSANDRO;DEPEND GEL IQ"). The brands actually
    # present are the ones read per row, so report those.
    brands_found = sorted({i["brand"] for i in result.items if i.get("brand")})
    if brands_found:
        result.brand = ", ".join(brands_found)

    return result


def _parse_feed(result, candidates, brand, country, banner, country3):
    """Parse one feed (one metadata group) into `result`."""
    ws, header_row1, header_row2, total_row = _pick_authoritative(candidates)

    weeks = _week_columns(ws, header_row1, header_row2, ws.max_column)
    if not weeks:
        raise ValueError("No year-week columns detected")
    # A week label appearing twice would make the second column silently
    # overwrite the first on upsert (same primary key) — surface it.
    week_labels = [w for w, _, _ in weeks]
    dupes = sorted({w for w in week_labels if week_labels.count(w) > 1})
    if dupes:
        result.warnings.append(
            f"Duplicate week column(s) in the sheet: {', '.join(dupes)} — "
            "later columns overwrite earlier ones for the same week."
        )
    first_week_col = weeks[0][1]
    attr_cols = _attribute_columns(ws, header_row1, header_row2, first_week_col)

    sku_col = attr_cols.get("SKU No.", 3)
    desc_col = attr_cols.get("Article Description")
    headgroup_col = attr_cols.get("Headgroup")
    size_col = attr_cols.get("Size")
    colour_col = attr_cols.get("Colour")
    type_col = attr_cols.get("Type")
    package_col = attr_cols.get("Content") or attr_cols.get("Package")
    price_col = attr_cols.get("Price") or attr_cols.get("Consumer")
    gtin_col = attr_cols.get("GTIN/PLU")
    # One export can cover several brands at once — then the metadata block
    # lists them joined with ';' ("ALESSANDRO;DEPEND GEL IQ") and only the
    # per-row "Brand" column says which brand a row belongs to. Using the
    # metadata value for every row would invent a brand named after the whole
    # list and detach the rows from that brand's own history.
    brand_col = attr_cols.get("Brand")

    seen_skus = set()
    duplicate_rows = 0
    non_numeric_cells = 0
    data_start = header_row2 + 1
    data_end = (total_row - 1) if total_row else ws.max_row
    # Where this feed's facts start, so the Total-row reconciliation below
    # sums only THIS feed and not other feeds already parsed from the file.
    facts_start = len(result.facts)

    for r in range(data_start, data_end + 1):
        sku = ws.cell(row=r, column=sku_col).value
        if sku is None:
            continue
        try:
            sku = str(int(sku))
        except (TypeError, ValueError):
            result.warnings.append(f"Row {r}: non-numeric SKU '{sku}', skipped.")
            continue

        row_brand = brand
        if brand_col:
            raw_brand = ws.cell(row=r, column=brand_col).value
            if raw_brand is not None and str(raw_brand).strip():
                row_brand = str(raw_brand).strip()

        if (row_brand, sku) in seen_skus:
            duplicate_rows += 1
            continue
        seen_skus.add((row_brand, sku))

        result.items.append(
            {
                "sku": sku,
                "brand": row_brand,
                "article_description": ws.cell(row=r, column=desc_col).value if desc_col else None,
                "headgroup": ws.cell(row=r, column=headgroup_col).value if headgroup_col else None,
                "size": ws.cell(row=r, column=size_col).value if size_col else None,
                "colour": ws.cell(row=r, column=colour_col).value if colour_col else None,
                "type": ws.cell(row=r, column=type_col).value if type_col else None,
                "package_content": ws.cell(row=r, column=package_col).value if package_col else None,
                "consumer_price": ws.cell(row=r, column=price_col).value if price_col else None,
                "gtin": str(ws.cell(row=r, column=gtin_col).value) if gtin_col and ws.cell(row=r, column=gtin_col).value else None,
            }
        )

        for year_week, vol_col, val_col in weeks:
            raw_vol = ws.cell(row=r, column=vol_col).value
            raw_val = ws.cell(row=r, column=val_col).value if val_col else None
            if raw_vol is None and raw_val is None:
                continue
            vol = _to_number(raw_vol)
            val = _to_number(raw_val)
            # Also counts a cell that's blank while its sibling has data (e.g.
            # volume reported, value not) — without this it was silently
            # booked as 0, indistinguishable from a real reported zero.
            if vol is None:
                non_numeric_cells += 1
            if val is None:
                non_numeric_cells += 1
            result.facts.append(
                {
                    "retailer": RETAILER_ID,
                    "brand": row_brand,
                    "country": country,
                    "banner": banner,
                    "unit": sku,
                    "period": year_week,
                    "sales_volume": vol if vol is not None else 0,
                    "sales_value": val if val is not None else 0,
                }
            )

    if duplicate_rows:
        result.warnings.append(
            f"Skipped {duplicate_rows} duplicate SKU row(s) found after the "
            "first occurrence (same SKU repeated, e.g. differing GTIN)."
        )
    if non_numeric_cells:
        result.warnings.append(
            f"{non_numeric_cells} volume/value cell(s) were blank or contained "
            "non-numeric text and were treated as 0 — check the source file "
            "before acting on this."
        )

    # Reconcile against the printed Total row per week, if present.
    if total_row:
        totals_by_week = {}
        for year_week, vol_col, val_col in weeks:
            totals_by_week[year_week] = ws.cell(row=total_row, column=vol_col).value

        computed = {}
        for f in result.facts[facts_start:]:
            computed[f["period"]] = computed.get(f["period"], 0) + (f["sales_volume"] or 0)

        mismatches = 0
        for year_week, expected_total in totals_by_week.items():
            if expected_total is None:
                continue
            got = computed.get(year_week, 0)
            if abs(got - expected_total) > 0.01:
                mismatches += 1
        if mismatches:
            result.warnings.append(
                f"{mismatches} week(s) had a sales-volume total that did not "
                "reconcile with the file's own Total row — review before trusting."
            )

    if not brand or not country3 or not banner:
        result.warnings.append(
            "Could not read Brand/Country/Formula from the metadata block "
            "(rows 1-7) — check the file layout."
        )
